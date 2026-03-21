"""
TEST FIXTURE ONLY — fake data created at owner's explicit direction for capability testing.

Generates a minimal Honda January 2026 bank rec XLSX with an "OS CKS" sheet.
The 18 "real" outstanding checks match the unmatched bank checks from Honda Feb
2026 testing (checks that cleared the bank in Feb but were cut in a prior period).

Usage:
    python tests/create_jan_rec_fixture.py
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, numbers

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "fixtures")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "HONDA_BANK_REC_0126_TEST.xlsx")

# ── The 18 checks that will clear in February ────────────────────────────
# These are the unmatched bank checks from the Honda Feb 2026 test run.
# They were outstanding as of January 31 and cleared in February.
REAL_OUTSTANDING = [
    (122651, 1522.50),
    (122665, 9000.00),
    (122673, 2000.00),
    (122683, 29000.00),
    (128395, 1143.75),
    (128436, 750.07),
    (128560, 1173.50),
    (128611, 499.00),
    (128675, 1077.50),
    (128710, 383.00),
    (128711, 336.51),
    (128808, 7.50),
    (128859, 1170.00),
    (128940, 152.00),
    (128961, 973.75),
    (129125, 92.00),
    (129139, 945.00),
    (129213, 65.76),
]

# ── ~20 fictional stale outstanding checks (older check numbers) ─────────
# These are checks that were outstanding as of Jan 31 but did NOT clear
# in February — they remain outstanding.  Lower check numbers, small amounts.
FICTIONAL_OUTSTANDING = [
    (108012, 45.00, "TEST VENDOR A"),
    (108345, 120.75, "TEST VENDOR B"),
    (109001, 88.50, "TEST VENDOR C"),
    (109234, 210.00, "TEST VENDOR D"),
    (109890, 15.25, "TEST VENDOR E"),
    (110022, 330.00, "TEST VENDOR F"),
    (110456, 62.99, "TEST VENDOR G"),
    (110789, 175.00, "TEST VENDOR H"),
    (111002, 99.00, "TEST VENDOR I"),
    (111345, 540.00, "TEST VENDOR J"),
    (112001, 28.50, "TEST VENDOR K"),
    (112234, 415.00, "TEST VENDOR L"),
    (112890, 67.25, "TEST VENDOR M"),
    (113022, 195.50, "TEST VENDOR N"),
    (113456, 82.00, "TEST VENDOR O"),
    (113789, 310.00, "TEST VENDOR P"),
    (114002, 47.75, "TEST VENDOR Q"),
    (114345, 155.00, "TEST VENDOR R"),
    (114890, 73.00, "TEST VENDOR S"),
    (115022, 260.00, "TEST VENDOR T"),
]


def create_fixture():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OS CKS"

    # ── Header row ────────────────────────────────────────────────────
    headers = {
        "A1": "Year",
        "B1": "Month",
        "C1": "Check#",
        "D1": "",
        "E1": "",
        "F1": "Amount",
        "G1": "",
        "H1": "Payee",
    }
    bold = Font(bold=True)
    for cell_ref, value in headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = bold

    # ── Column widths ─────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["H"].width = 30

    # ── Helper to write a check row ───────────────────────────────────
    running_total = 0.0

    def write_check(row_num, year, month, check_num, amount, payee):
        nonlocal running_total
        if year:
            ws.cell(row=row_num, column=1, value=year)
        if month:
            ws.cell(row=row_num, column=2, value=month)
        ws.cell(row=row_num, column=3, value=check_num)
        amt_cell = ws.cell(row=row_num, column=6, value=amount)
        amt_cell.number_format = '#,##0.00'
        ws.cell(row=row_num, column=8, value=payee)
        running_total += amount

    # ── Write fictional stale checks (2025 vintage) ──────────────────
    row = 2
    write_check(row, 2025, 6, FICTIONAL_OUTSTANDING[0][0],
                FICTIONAL_OUTSTANDING[0][1], FICTIONAL_OUTSTANDING[0][2])
    row += 1

    for ck_num, amt, payee in FICTIONAL_OUTSTANDING[1:5]:
        write_check(row, None, None, ck_num, amt, payee)
        row += 1

    # New month group
    write_check(row, None, 7, FICTIONAL_OUTSTANDING[5][0],
                FICTIONAL_OUTSTANDING[5][1], FICTIONAL_OUTSTANDING[5][2])
    row += 1

    for ck_num, amt, payee in FICTIONAL_OUTSTANDING[6:10]:
        write_check(row, None, None, ck_num, amt, payee)
        row += 1

    # Another month group
    write_check(row, None, 8, FICTIONAL_OUTSTANDING[10][0],
                FICTIONAL_OUTSTANDING[10][1], FICTIONAL_OUTSTANDING[10][2])
    row += 1

    for ck_num, amt, payee in FICTIONAL_OUTSTANDING[11:15]:
        write_check(row, None, None, ck_num, amt, payee)
        row += 1

    # Another month group
    write_check(row, None, 9, FICTIONAL_OUTSTANDING[15][0],
                FICTIONAL_OUTSTANDING[15][1], FICTIONAL_OUTSTANDING[15][2])
    row += 1

    for ck_num, amt, payee in FICTIONAL_OUTSTANDING[16:]:
        write_check(row, None, None, ck_num, amt, payee)
        row += 1

    # ── Write the 18 real outstanding checks (2025-2026 vintage) ─────
    # Year=2025, month=11 for the first batch, then 12, then 2026/1
    write_check(row, 2025, 11, REAL_OUTSTANDING[0][0],
                REAL_OUTSTANDING[0][1], "PRIOR PERIOD TEST")
    row += 1
    for ck_num, amt in REAL_OUTSTANDING[1:4]:
        write_check(row, None, None, ck_num, amt, "PRIOR PERIOD TEST")
        row += 1

    write_check(row, None, 12, REAL_OUTSTANDING[4][0],
                REAL_OUTSTANDING[4][1], "PRIOR PERIOD TEST")
    row += 1
    for ck_num, amt in REAL_OUTSTANDING[5:10]:
        write_check(row, None, None, ck_num, amt, "PRIOR PERIOD TEST")
        row += 1

    write_check(row, 2026, 1, REAL_OUTSTANDING[10][0],
                REAL_OUTSTANDING[10][1], "PRIOR PERIOD TEST")
    row += 1
    for ck_num, amt in REAL_OUTSTANDING[11:]:
        write_check(row, None, None, ck_num, amt, "PRIOR PERIOD TEST")
        row += 1

    # ── Total row ─────────────────────────────────────────────────────
    row += 1  # blank separator
    total_cell = ws.cell(row=row, column=6, value=running_total)
    total_cell.number_format = '#,##0.00'
    total_cell.font = bold
    ws.cell(row=row, column=8, value="TOTAL OUTSTANDING").font = bold

    wb.save(OUTPUT_FILE)
    print(f"Created fixture: {OUTPUT_FILE}")
    print(f"  {len(FICTIONAL_OUTSTANDING)} fictional stale checks")
    print(f"  {len(REAL_OUTSTANDING)} real outstanding checks (will clear in Feb)")
    print(f"  Total outstanding: ${running_total:,.2f}")
    print(f"  Total rows (excl header/total): {len(FICTIONAL_OUTSTANDING) + len(REAL_OUTSTANDING)}")


if __name__ == "__main__":
    create_fixture()
