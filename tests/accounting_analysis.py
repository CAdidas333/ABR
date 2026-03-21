"""
Accounting Analysis — Honda Feb 2026

Senior accounting analyst review of ABR engine results against Jay's completed
bank rec (ground truth). Every dollar must be accounted for.

Ground truth from HONDA BANK REC 0226.xlsx:
  - Truist Statement Balance: $2,845,158.77
  - Columbia/Fulton Bank Balance: $27,906.73
  - Sweep: $2,172,826.11
  - O/S Sweep Entry: -$103,943.60
  - Outstanding deposits: $274,213.56 (6 items)
  - Outstanding checks: -$528,506.53 (661 items)
  - Outstanding PR checks: -$6,950.99 (12 items)
  - Adjusted bank balance: $4,680,704.05
  - GL Balance Acct 202: $4,789,019.01
  - Fulton Bank Fee: -$215.95
  - Outstanding Journal Entries: -$108,097.26 (13 items)
  - TO BAL: -$1.75
  - Adjusted book balance: $4,680,704.05
  - Difference: $0.00
"""

import sys
import os
import csv
from collections import defaultdict
from datetime import date

sys.path.insert(0, os.path.dirname(__file__))

from matching_engine import (
    parse_bofa_bai_csv,
    parse_dms_xlsx,
    run_full_reconciliation,
    MatchConfig,
    Transaction,
)

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. pip install openpyxl")
    sys.exit(1)


# ============================================================================
# File paths
# ============================================================================
DOCS = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'docs')
BANK_FILE = os.path.join(DOCS, 'Honda Bkstmt.csv')
FEB_GL_FILE = os.path.join(DOCS, 'honda Feb GL.xlsx')
JAN_GL_FILE = os.path.join(DOCS, 'honda Jan GL.xlsx')
REC_FILE = os.path.join(DOCS, 'HONDA BANK REC 0226.xlsx')


# ============================================================================
# Helpers
# ============================================================================
def fmt(amount):
    """Format a dollar amount."""
    if amount < 0:
        return f"-${abs(amount):,.2f}"
    return f"${amount:,.2f}"


def load_bank_rec_data():
    """Load all ground-truth data from Jay's completed bank rec."""
    wb = openpyxl.load_workbook(REC_FILE, data_only=True)

    # --- RECONCILIATION sheet ---
    ws_rec = wb['RECONCILLIATION']
    rec_data = {}
    for row in ws_rec.iter_rows(min_row=1, max_row=ws_rec.max_row, values_only=True):
        # Store all non-empty rows for inspection
        vals = [v for v in row if v is not None]
        if vals:
            rec_data[str(vals)] = row

    # --- OS CKS sheet ---
    ws_cks = wb['OS CKS']
    os_checks = []
    # Diagnostic: print first few data rows to identify column layout
    diag_rows = []
    for row_idx, row in enumerate(ws_cks.iter_rows(min_row=5, values_only=True)):
        if row_idx < 5:
            diag_rows.append(list(row))

        # Columns: A=Year, B=Month, C=Check#, D=?, E=Amount, F=Payee...
        if row[2] is not None:
            ck_num = str(int(row[2])) if isinstance(row[2], (int, float)) else str(row[2]).strip()
            # Try to find the amount -- check multiple columns
            amount = 0.0
            payee = ''
            for col_idx in [4, 3, 5, 6, 7]:
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None and isinstance(val, (int, float)) and val != 0:
                    amount = float(val)
                    break
            # Payee: look for first string after the check number
            for col_idx in [5, 6, 7, 8]:
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None and isinstance(val, str) and val.strip():
                    payee = val.strip()
                    break
            year = row[0]
            month = str(row[1]).strip() if row[1] else ''
            if ck_num:
                os_checks.append({
                    'check_number': ck_num,
                    'amount': amount,
                    'payee': payee,
                    'year': year,
                    'month': month,
                })

    # Print diagnostic info
    print(f"  [DIAG] OS CKS first 5 data rows (starting row 5):")
    for i, r in enumerate(diag_rows):
        print(f"    Row {5+i}: {r[:10]}")

    # --- OS PR CKS sheet ---
    ws_pr = wb['OS PR CKS']
    os_pr_checks = []
    diag_pr = []
    for row_idx, row in enumerate(ws_pr.iter_rows(min_row=5, values_only=True)):
        if row_idx < 3:
            diag_pr.append(list(row))
        if row[2] is not None:
            ck_num = str(int(row[2])) if isinstance(row[2], (int, float)) else str(row[2]).strip()
            # Try to find the amount
            amount = 0.0
            for col_idx in [4, 3, 5]:
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None and isinstance(val, (int, float)) and val != 0:
                    amount = float(val)
                    break
            payee = str(row[5]).strip() if len(row) > 5 and row[5] else ''
            if ck_num:
                os_pr_checks.append({
                    'check_number': ck_num,
                    'amount': amount,
                    'payee': payee,
                })
    print(f"  [DIAG] OS PR CKS first 3 data rows:")
    for i, r in enumerate(diag_pr):
        print(f"    Row {5+i}: {r[:8]}")

    # --- OS DEPOSITS sheet ---
    ws_dep = wb['OS DEPOSITS']
    os_deposits = []
    for row in ws_dep.iter_rows(min_row=4, values_only=True):
        if row[3] is not None and row[4] is not None:
            ref = str(row[3]).strip()
            amount = float(row[4])
            if ref and amount != 0:
                os_deposits.append({
                    'ref': ref,
                    'amount': amount,
                })

    wb.close()
    return os_checks, os_pr_checks, os_deposits


def load_rec_journal_entries():
    """Load outstanding journal entries from the RECONCILIATION sheet."""
    wb = openpyxl.load_workbook(REC_FILE, data_only=True)
    ws = wb['RECONCILLIATION']

    journal_entries = []
    in_je_section = False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        # Look for the journal entries section
        row_vals = [str(v).strip() if v is not None else '' for v in row]
        row_str = ' '.join(row_vals).upper()

        if 'JOURNAL' in row_str and 'ENTRIES' in row_str:
            in_je_section = True
            continue

        if in_je_section:
            # Look for rows with amounts
            amounts = []
            refs = []
            for v in row:
                if v is None:
                    continue
                if isinstance(v, (int, float)):
                    amounts.append(v)
                elif isinstance(v, str) and v.strip():
                    refs.append(v.strip())

            if amounts:
                journal_entries.append({
                    'refs': refs,
                    'amounts': amounts,
                    'raw': row,
                })

            # Stop if we hit a blank row or totals row
            all_none = all(v is None for v in row)
            if all_none and len(journal_entries) > 0:
                break
            if any(str(v).upper().strip() in ('TOTAL', 'TO BAL') for v in row if v is not None):
                break

    wb.close()
    return journal_entries


def load_full_rec_sheet():
    """Load every row from the RECONCILIATION sheet for bridge building."""
    wb = openpyxl.load_workbook(REC_FILE, data_only=True)
    ws = wb['RECONCILLIATION']
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        rows.append(row)
    wb.close()
    return rows


# ============================================================================
# Main Analysis
# ============================================================================
def main():
    print("=" * 80)
    print("  ACCOUNTING ANALYSIS — Honda Feb 2026 Bank Reconciliation")
    print("  Analyst Review of ABR Engine vs Jay's Ground Truth")
    print("=" * 80)

    # ------------------------------------------------------------------
    # Step 0: Load everything
    # ------------------------------------------------------------------
    print("\n[LOADING DATA]")
    bank_txns = parse_bofa_bai_csv(BANK_FILE)
    print(f"  Bank transactions loaded: {len(bank_txns)}")

    feb_dms = parse_dms_xlsx(FEB_GL_FILE)
    print(f"  Feb GL transactions: {len(feb_dms)}")

    jan_dms = parse_dms_xlsx(JAN_GL_FILE)
    print(f"  Jan GL transactions: {len(jan_dms)}")

    os_checks, os_pr_checks, os_deposits = load_bank_rec_data()
    print(f"  Jay's outstanding AP checks: {len(os_checks)}")
    print(f"  Jay's outstanding PR checks: {len(os_pr_checks)}")
    print(f"  Jay's outstanding deposits: {len(os_deposits)}")

    rec_rows = load_full_rec_sheet()
    print(f"  Reconciliation sheet rows: {len(rec_rows)}")

    # Run ABR engine
    print("\n[RUNNING ABR ENGINE]")
    config = MatchConfig()
    results = run_full_reconciliation(bank_txns, feb_dms, config, prior_dms_txns=jan_dms)

    total_matches = len(results['all_matches'])
    unmatched_bank = results['unmatched_bank']
    unmatched_dms = results['unmatched_dms']

    print(f"  Total matches: {total_matches}")
    print(f"  Unmatched bank: {len(unmatched_bank)}")
    print(f"  Unmatched DMS:  {len(unmatched_dms)}")

    # ======================================================================
    # TASK 1: FULL RECONCILIATION BRIDGE
    # ======================================================================
    print("\n" + "=" * 80)
    print("  TASK 1: FULL RECONCILIATION BRIDGE")
    print("=" * 80)

    # --- Print Jay's reconciliation sheet line by line ---
    print("\n  Jay's Reconciliation Sheet (raw):")
    print("  " + "-" * 76)
    for i, row in enumerate(rec_rows, 1):
        vals = [v for v in row if v is not None]
        if vals:
            # Format nicely
            formatted = []
            for v in row:
                if v is None:
                    continue
                if isinstance(v, float):
                    formatted.append(fmt(v))
                else:
                    formatted.append(str(v))
            print(f"    Row {i:2d}: {' | '.join(formatted)}")

    # --- Bank Statement Summary ---
    print("\n  BRIDGE: Bank Statement -> Adjusted Bank Balance")
    print("  " + "-" * 76)

    bank_debits = sum(t.amount for t in bank_txns if t.amount < 0)
    bank_credits = sum(t.amount for t in bank_txns if t.amount > 0)
    bank_net = sum(t.amount for t in bank_txns)

    # Break down by type
    bank_by_type = defaultdict(lambda: {'count': 0, 'total': 0.0})
    for t in bank_txns:
        bank_by_type[t.type_code]['count'] += 1
        bank_by_type[t.type_code]['total'] += t.amount

    print(f"    Bank statement transaction breakdown:")
    for tc in sorted(bank_by_type.keys()):
        d = bank_by_type[tc]
        print(f"      {tc:10s}: {d['count']:4d} txns  {fmt(d['total']):>15s}")
    print(f"      {'TOTAL':10s}: {len(bank_txns):4d} txns  {fmt(bank_net):>15s}")

    # Sweep analysis
    sweeps = [t for t in bank_txns if t.type_code == 'SWEEP']
    sweep_total = sum(t.amount for t in sweeps)
    print(f"\n    Sweep entries: {len(sweeps)} transactions, net {fmt(sweep_total)}")
    for s in sorted(sweeps, key=lambda x: x.transaction_date):
        print(f"      {s.transaction_date}  {fmt(s.amount):>15s}  {s.description[:50]}")

    # Securities (233)
    securities = [t for t in bank_txns if t.type_code == 'SEC']
    sec_total = sum(t.amount for t in securities)
    print(f"\n    Securities entries: {len(securities)} transactions, net {fmt(sec_total)}")
    for s in securities:
        print(f"      {s.transaction_date}  {fmt(s.amount):>15s}  {s.description[:50]}")

    # Jay's rec has:
    #   Truist Statement Balance:  $2,845,158.77
    #   Columbia/Fulton Bank:      $27,906.73
    #   Sweep:                     $2,172,826.11
    #   O/S Sweep Entry:           -$103,943.60
    # These form the opening position BEFORE bank activity

    # --- DMS GL Summary ---
    print(f"\n  DMS Feb GL breakdown:")
    dms_by_type = defaultdict(lambda: {'count': 0, 'total': 0.0})
    for t in feb_dms:
        dms_by_type[t.type_code]['count'] += 1
        dms_by_type[t.type_code]['total'] += t.amount

    dms_net = sum(t.amount for t in feb_dms)
    for tc in sorted(dms_by_type.keys()):
        d = dms_by_type[tc]
        print(f"      {tc:10s}: {d['count']:4d} txns  {fmt(d['total']):>15s}")
    print(f"      {'TOTAL':10s}: {len(feb_dms):4d} txns  {fmt(dms_net):>15s}")

    # SRC=9 balance forward
    balfwd = [t for t in feb_dms if t.type_code == '9' or 'BALFWD' in t.reference_number.upper()]
    if balfwd:
        print(f"\n    SRC=9 / BALFWD entries: {len(balfwd)}")
        for b in balfwd:
            print(f"      ID={b.transaction_id} Ref={b.reference_number} "
                  f"{b.transaction_date} {fmt(b.amount)} {b.description[:60]}")

    # ======================================================================
    # TASK 2: VALIDATE OUTSTANDING CHECKS OVERLAP
    # ======================================================================
    print("\n" + "=" * 80)
    print("  TASK 2: OUTSTANDING CHECKS OVERLAP VALIDATION")
    print("=" * 80)

    jay_check_nums = {c['check_number'] for c in os_checks}
    jay_pr_check_nums = {c['check_number'] for c in os_pr_checks}
    all_jay_os = jay_check_nums | jay_pr_check_nums

    jay_os_total = sum(c['amount'] for c in os_checks)
    jay_pr_total = sum(c['amount'] for c in os_pr_checks)

    print(f"\n  Jay's outstanding checks:")
    print(f"    AP checks: {len(os_checks)} totaling {fmt(jay_os_total)}")
    print(f"    PR checks: {len(os_pr_checks)} totaling {fmt(jay_pr_total)}")
    print(f"    Combined:  {len(all_jay_os)} unique check numbers")

    # Bank statement check numbers
    bank_check_set = {t.check_number for t in bank_txns if t.check_number}
    print(f"\n  Bank statement unique check numbers: {len(bank_check_set)}")

    # Sanity check: no check should be on BOTH the bank statement and outstanding list
    cleared_but_os = bank_check_set & all_jay_os
    if cleared_but_os:
        print(f"  WARNING: {len(cleared_but_os)} checks on BOTH bank stmt and outstanding list!")
        for ck in sorted(cleared_but_os)[:20]:
            print(f"    Check #{ck}")
    else:
        print(f"  OK: No checks on both bank statement and outstanding list")

    # DMS Feb GL check numbers (SRC=6 / CHK type)
    feb_dms_checks = {t.reference_number: t for t in feb_dms if t.type_code == 'CHK'}
    feb_dms_check_nums = set()
    for t in feb_dms:
        if t.type_code == 'CHK' and t.reference_number:
            # Strip trailing letters
            ref = t.reference_number
            while ref and ref[-1].isalpha():
                ref = ref[:-1]
            if ref:
                feb_dms_check_nums.add(ref)
    print(f"\n  Feb DMS check numbers (CHK/SRC=6): {len(feb_dms_check_nums)}")

    # DMS Jan GL check numbers
    jan_dms_check_nums = set()
    for t in jan_dms:
        if t.type_code == 'CHK' and t.reference_number:
            ref = t.reference_number
            while ref and ref[-1].isalpha():
                ref = ref[:-1]
            if ref:
                jan_dms_check_nums.add(ref)
    print(f"  Jan DMS check numbers (CHK/SRC=6): {len(jan_dms_check_nums)}")

    # Our unmatched DMS checks from Feb
    unmatched_feb_dms_checks = [t for t in unmatched_dms if t.type_code == 'CHK']
    unmatched_feb_check_nums = set()
    for t in unmatched_feb_dms_checks:
        ref = t.reference_number
        while ref and ref[-1].isalpha():
            ref = ref[:-1]
        if ref:
            unmatched_feb_check_nums.add(ref)
    print(f"  Our unmatched Feb DMS checks: {len(unmatched_feb_dms_checks)} "
          f"({len(unmatched_feb_check_nums)} unique check nums)")

    # Overlap analysis
    jay_in_our_feb_unmatched = jay_check_nums & unmatched_feb_check_nums
    jay_in_feb_dms = jay_check_nums & feb_dms_check_nums
    jay_in_jan_dms = jay_check_nums & jan_dms_check_nums
    jay_not_in_any_dms = jay_check_nums - feb_dms_check_nums - jan_dms_check_nums

    print(f"\n  Jay's {len(jay_check_nums)} outstanding AP checks overlap:")
    print(f"    In Feb DMS GL:           {len(jay_in_feb_dms)}")
    print(f"    In Jan DMS GL:           {len(jay_in_jan_dms)}")
    print(f"    In our unmatched Feb DMS: {len(jay_in_our_feb_unmatched)}")
    print(f"    NOT in Feb or Jan DMS:   {len(jay_not_in_any_dms)} (older months)")

    # Quantify the older-month checks
    if jay_not_in_any_dms:
        old_checks = [c for c in os_checks if c['check_number'] in jay_not_in_any_dms]
        old_total = sum(c['amount'] for c in old_checks)
        print(f"    Total $ of older-month checks: {fmt(old_total)}")

        # Show year/month breakdown
        by_period = defaultdict(lambda: {'count': 0, 'total': 0.0})
        for c in old_checks:
            period = f"{c['year']}-{c['month']}" if c['year'] and c['month'] else "unknown"
            by_period[period]['count'] += 1
            by_period[period]['total'] += c['amount']
        print(f"    By period:")
        for period in sorted(by_period.keys()):
            d = by_period[period]
            print(f"      {period:15s}: {d['count']:4d} checks  {fmt(d['total']):>12s}")

    # Checks in our unmatched Feb DMS that are NOT on Jay's outstanding list
    # (these are checks that cleared the bank in Feb -- they should be matched)
    our_unmatched_not_on_jay = unmatched_feb_check_nums - all_jay_os
    if our_unmatched_not_on_jay:
        print(f"\n  Our unmatched Feb DMS checks NOT on Jay's O/S list: {len(our_unmatched_not_on_jay)}")
        print(f"  (These should have matched to bank -- investigate)")
        for ck in sorted(our_unmatched_not_on_jay)[:10]:
            txn = next((t for t in unmatched_feb_dms_checks
                       if t.reference_number.rstrip('ABCDEFGHIJKLMNOPQRSTUVWXYZ') == ck), None)
            if txn:
                print(f"    Check #{ck}  {fmt(txn.amount):>12s}  {txn.transaction_date}  {txn.description[:40]}")

    # ======================================================================
    # TASK 3: INVESTIGATE 42 UNMATCHED FINDEP (SRC=11) DMS ITEMS
    # ======================================================================
    print("\n" + "=" * 80)
    print("  TASK 3: UNMATCHED FINDEP (SRC=11) DMS ITEMS")
    print("=" * 80)

    unmatched_findep = [t for t in unmatched_dms if t.type_code == 'FINDEP']
    findep_total = sum(t.amount for t in unmatched_findep)
    print(f"\n  Unmatched FINDEP items: {len(unmatched_findep)}")
    print(f"  Net total: {fmt(findep_total)}")

    # Categorize by description keywords
    categories = defaultdict(lambda: {'items': [], 'total': 0.0})
    for t in unmatched_findep:
        desc_upper = t.description.upper()
        ref_upper = t.reference_number.upper()

        if 'FULTON' in desc_upper or 'COLUMBIA' in desc_upper:
            cat = 'Fulton/Columbia Bank'
        elif 'AHM' in desc_upper or 'HONDA' in desc_upper or 'AHFC' in desc_upper:
            cat = 'AHM/Honda'
        elif 'ACV' in desc_upper:
            cat = 'ACV Auctions'
        elif 'SWEEP' in desc_upper or 'DATASCAN' in desc_upper:
            cat = 'Sweep/DataScan'
        elif 'MERCHANT' in desc_upper or 'WORLDPAY' in desc_upper or 'BANKCARD' in desc_upper:
            cat = 'Merchant Processing'
        elif 'ZURICH' in desc_upper:
            cat = 'Zurich Insurance'
        elif 'TITLE' in desc_upper or 'TTI' in desc_upper or 'HFE' in desc_upper:
            cat = 'Title/HFE'
        elif 'FEE' in desc_upper:
            cat = 'Fees'
        elif 'BALFWD' in ref_upper or ref_upper.startswith('BALFWD'):
            cat = 'Balance Forward'
        else:
            cat = 'Other'

        categories[cat]['items'].append(t)
        categories[cat]['total'] += t.amount

    print(f"\n  Categorized breakdown:")
    for cat in sorted(categories.keys()):
        d = categories[cat]
        print(f"    {cat:25s}: {len(d['items']):3d} items  {fmt(d['total']):>15s}")
        for t in sorted(d['items'], key=lambda x: abs(x.amount), reverse=True)[:5]:
            print(f"      ID={t.transaction_id:4d} Ref={t.reference_number:15s} "
                  f"{t.transaction_date} {fmt(t.amount):>12s} {t.description[:50]}")
        if len(d['items']) > 5:
            print(f"      ... and {len(d['items']) - 5} more")

    # Check for self-canceling pairs that Phase -1 might have missed
    print(f"\n  Self-canceling pair analysis (FINDEP):")
    potential_pairs = []
    findep_sorted = sorted(unmatched_findep, key=lambda x: abs(x.amount))
    for i, a in enumerate(findep_sorted):
        for j in range(i + 1, len(findep_sorted)):
            b = findep_sorted[j]
            if abs(a.amount + b.amount) < 0.01 and a.amount != 0:
                potential_pairs.append((a, b))

    if potential_pairs:
        print(f"    Found {len(potential_pairs)} potential self-canceling pairs:")
        for a, b in potential_pairs:
            print(f"      {fmt(a.amount):>12s} (Ref={a.reference_number}, {a.transaction_date}) + "
                  f"{fmt(b.amount):>12s} (Ref={b.reference_number}, {b.transaction_date})")
            print(f"        A desc: {a.description[:60]}")
            print(f"        B desc: {b.description[:60]}")
    else:
        print(f"    No self-canceling pairs found in unmatched FINDEP items")

    # Check: do any correspond to Jay's 13 outstanding journal entries?
    print(f"\n  Mapping to Jay's outstanding journal entries:")
    # Load journal entries from rec sheet
    je_entries = load_rec_journal_entries()
    if je_entries:
        print(f"    Found {len(je_entries)} journal entry rows in rec sheet:")
        for je in je_entries:
            print(f"      {je['refs']}  amounts={je['amounts']}")
    else:
        print(f"    Could not parse journal entries from rec sheet -- checking amounts manually")

    # Map Jay's 13 JEs to our unmatched FINDEP by reference number
    # Jay's JE references from the rec sheet:
    jays_jes = {
        '022726AHM':     111480.44,
        '022826AHM':      73014.47,   # Jay wrote "02268626AHM" (typo) -- it's 022826AHM
        'FLRPLN0226':     -8289.84,
        'HFS022726':     -27301.33,
        'TTEC022726':    -23489.98,
        'TTEC022726_B':    -150.50,   # Second TTEC022726 entry (TTTEC022726 in DMS)
        'ZCERT022726':    -2457.00,
        'ZLWST022726':     -188.00,
        'ZS022526':        -260.00,   # ZS021826 in DMS (date variant)
        'ZS022726':        -260.00,   # Jay wrote "ZS02272+6" (typo)
        'ZSCRP022726':    -1139.00,
        'ZVSC022726':     -1130.00,
        'ZVSC022726A':   -11732.00,
    }
    jays_je_total = sum(jays_jes.values())

    print(f"\n    Mapping Jay's 13 JEs to our unmatched FINDEP:")
    print(f"    Jay's JE total: {fmt(jays_je_total)}")
    mapped_count = 0
    mapped_total = 0.0
    used_ids = set()  # Track already-mapped FINDEP IDs to avoid double-mapping
    for je_ref, je_amt in jays_jes.items():
        # Find matching FINDEP by reference + amount (may need fuzzy match)
        match = None

        # Special cases first (before generic match)
        if je_ref == 'TTEC022726_B':
            # This is the $-150.50 entry -- maps to TTTEC022726 (note triple-T)
            for t in unmatched_findep:
                if t.transaction_id not in used_ids and 'TTTEC022726' in t.reference_number.upper():
                    match = t
                    break
        elif je_ref == 'ZS022526':
            for t in unmatched_findep:
                if t.transaction_id not in used_ids and t.reference_number.upper() == 'ZS021826':
                    match = t
                    break
        else:
            # Generic match by reference, prefer amount match when multiple refs exist
            je_ref_upper = je_ref.upper()
            candidates = [t for t in unmatched_findep
                          if t.transaction_id not in used_ids
                          and t.reference_number.upper() == je_ref_upper]
            if len(candidates) == 1:
                match = candidates[0]
            elif len(candidates) > 1:
                # Pick the one closest to Jay's amount
                candidates.sort(key=lambda t: abs(t.amount - je_amt))
                match = candidates[0]

        if match:
            used_ids.add(match.transaction_id)
        if match:
            mapped_count += 1
            mapped_total += match.amount
            status = "FOUND"
            diff = match.amount - je_amt
            diff_str = f" diff={fmt(diff)}" if abs(diff) > 0.01 else ""
            print(f"      {je_ref:18s} Jay={fmt(je_amt):>12s}  DMS={fmt(match.amount):>12s}  "
                  f"Ref={match.reference_number}{diff_str}")
        else:
            print(f"      {je_ref:18s} Jay={fmt(je_amt):>12s}  ** NOT FOUND IN UNMATCHED FINDEP **")

    print(f"\n    Mapped: {mapped_count}/13 JEs, DMS total: {fmt(mapped_total)}")
    print(f"    Jay's JE total:         {fmt(jays_je_total)}")
    print(f"    Our unmatched FINDEP:   {fmt(findep_total)}")
    print(f"    FINDEP minus JEs:       {fmt(findep_total - jays_je_total)}")
    print(f"    (The difference = FINDEP items that are NOT Jay's JEs, i.e.,")
    print(f"     self-canceling pairs + sweeps + other matched-elsewhere items)")

    # ALL unmatched FINDEP items listed
    print(f"\n  All {len(unmatched_findep)} unmatched FINDEP items (sorted by amount):")
    for t in sorted(unmatched_findep, key=lambda x: x.amount):
        print(f"    ID={t.transaction_id:4d} {t.transaction_date} Ref={t.reference_number:15s} "
              f"{fmt(t.amount):>12s}  {t.description[:55]}")

    # ======================================================================
    # TASK 4: INVESTIGATE 5 UNMATCHED ACH BANK ITEMS
    # ======================================================================
    print("\n" + "=" * 80)
    print("  TASK 4: UNMATCHED ACH BANK ITEMS")
    print("=" * 80)

    unmatched_ach = [t for t in unmatched_bank if 'ACH' in t.type_code]
    ach_credits = [t for t in unmatched_ach if t.amount > 0]
    ach_debits = [t for t in unmatched_ach if t.amount < 0]

    print(f"\n  Unmatched ACH items: {len(unmatched_ach)}")
    print(f"    Credits: {len(ach_credits)} totaling {fmt(sum(t.amount for t in ach_credits))}")
    print(f"    Debits:  {len(ach_debits)} totaling {fmt(sum(t.amount for t in ach_debits))}")

    # Read full transaction detail from the CSV for these items
    # We need to re-read the CSV to get the Transaction Detail column
    bank_detail_map = {}
    with open(BANK_FILE, 'r', newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader, 1):
            detail = row.get('Transaction Detail', '').strip()
            amount_str = row.get('Amount', '').strip()
            try:
                amount = float(amount_str)
            except ValueError:
                continue
            bank_detail_map[i] = {
                'amount': amount,
                'detail': detail,
                'description': row.get('Transaction Description', '').strip(),
                'type': row.get('Type', '').strip(),
                'bai_code': row.get('BAI Code', '').strip(),
                'date': row.get('As of Date', '').strip(),
                'customer_ref': row.get('Customer Reference', '').strip(),
            }

    print(f"\n  Detailed analysis of each unmatched ACH item:")
    for t in sorted(unmatched_ach, key=lambda x: x.amount):
        print(f"\n    Bank ID={t.transaction_id}: {t.transaction_date} {fmt(t.amount)}")
        print(f"      Type: {t.type_code}")
        print(f"      Description: {t.description[:80]}")

        # Get detail from CSV
        detail_info = bank_detail_map.get(t.transaction_id)
        if detail_info:
            detail_text = detail_info['detail']
            # Parse out key info from ACH detail
            if 'Payor Name:' in detail_text:
                payor_start = detail_text.index('Payor Name:') + len('Payor Name:')
                payor_end = detail_text.index('Payor ID:') if 'Payor ID:' in detail_text else payor_start + 30
                payor = detail_text[payor_start:payor_end].strip()
            else:
                payor = 'N/A'

            if 'Description:' in detail_text:
                desc_start = detail_text.index('Description:') + len('Description:')
                # Find next field
                for field in ['Payor Name:', 'Originating Bank:']:
                    if field in detail_text[desc_start:]:
                        desc_end = detail_text.index(field, desc_start)
                        break
                else:
                    desc_end = desc_start + 40
                ach_desc = detail_text[desc_start:desc_end].strip()
            else:
                ach_desc = 'N/A'

            if 'CUSTOMER ID' in detail_text:
                cust_start = detail_text.index('CUSTOMER ID') + len('CUSTOMER ID')
                cust_end = min(cust_start + 25, len(detail_text))
                cust_id = detail_text[cust_start:cust_end].strip().split()[0] if detail_text[cust_start:cust_end].strip() else 'N/A'
            else:
                cust_id = 'N/A'

            print(f"      ACH Payor: {payor}")
            print(f"      ACH Desc:  {ach_desc}")
            print(f"      Customer ID: {cust_id}")

        # Search for near-amount DMS matches
        near_matches = []
        for d in feb_dms:
            diff = abs(t.amount - d.amount)
            if diff > 0 and diff <= abs(t.amount) * 0.05 and diff < 5000:  # within 5%
                near_matches.append((d, diff))
        near_matches.sort(key=lambda x: x[1])

        if near_matches:
            print(f"      Near-amount DMS candidates (within 5%):")
            for d, diff in near_matches[:5]:
                print(f"        DMS ID={d.transaction_id} {d.transaction_date} "
                      f"{fmt(d.amount):>12s} diff={fmt(diff):>10s} "
                      f"Ref={d.reference_number} {d.description[:40]}")
        else:
            print(f"      No near-amount DMS candidates found")

    # ======================================================================
    # TASK 5: CALCULATE OUTSTANDING CHECKS FROM OUR DATA
    # ======================================================================
    print("\n" + "=" * 80)
    print("  TASK 5: OUTSTANDING CHECKS CALCULATION")
    print("=" * 80)

    # Our unmatched Feb DMS checks = checks written in Feb that didn't clear the bank
    our_os_checks = [t for t in unmatched_dms if t.type_code == 'CHK']
    our_os_check_total = sum(t.amount for t in our_os_checks)
    # DMS checks are negative (credits/outflows). For comparison with Jay, use abs value.
    our_os_check_abs = abs(our_os_check_total)

    print(f"\n  Our unmatched Feb DMS checks (SRC=6):")
    print(f"    Count: {len(our_os_checks)}")
    print(f"    Total (DMS sign): {fmt(our_os_check_total)}")
    print(f"    Total (absolute): {fmt(our_os_check_abs)}")

    # Note: Jay stores check amounts as positive on the O/S list.
    # The rec then shows "LESS O/S CHECKS" as a deduction.
    print(f"\n  Jay's outstanding check totals (positive = amount of check):")
    print(f"    AP checks: {len(os_checks)} totaling {fmt(jay_os_total)}")
    print(f"    PR checks: {len(os_pr_checks)} totaling {fmt(jay_pr_total)}")
    print(f"    Combined:  {fmt(jay_os_total + jay_pr_total)}")
    print(f"    Jay's rec shows: -$528,506.53 (LESS O/S CHECKS)")

    print(f"\n  Direct comparison (absolute values):")
    print(f"    Jay's AP O/S total:         {fmt(jay_os_total)}")
    print(f"    Our Feb unmatched CHK abs:   {fmt(our_os_check_abs)}")
    print(f"    Shortfall:                   {fmt(jay_os_total - our_os_check_abs)}")
    print(f"    (Shortfall = prior-period O/S checks not in Feb GL)")

    # The difference should be explained by:
    # 1. Checks from prior months still outstanding (on Jay's list but not in Feb DMS)
    # 2. Checks from Feb that cleared (matched) -- should not be on Jay's list
    # Jay's list includes ALL outstanding, including very old ones

    # Breakdown: Jay's O/S checks that ARE in Feb DMS vs not
    jay_in_feb = [c for c in os_checks
                  if c['check_number'] in feb_dms_check_nums]
    jay_not_in_feb = [c for c in os_checks
                      if c['check_number'] not in feb_dms_check_nums]

    jay_in_feb_total = sum(c['amount'] for c in jay_in_feb)
    jay_not_in_feb_total = sum(c['amount'] for c in jay_not_in_feb)

    print(f"\n  Jay's O/S checks breakdown:")
    print(f"    In Feb DMS GL:     {len(jay_in_feb)} checks  {fmt(jay_in_feb_total)}")
    print(f"    Not in Feb DMS GL: {len(jay_not_in_feb)} checks  {fmt(jay_not_in_feb_total)}")

    # For Jay's checks that ARE in Feb DMS, check if they were matched by our engine
    jay_in_feb_matched = []
    jay_in_feb_unmatched = []
    for c in jay_in_feb:
        ck = c['check_number']
        # Find the DMS transaction
        dms_txn = None
        for t in feb_dms:
            if t.type_code == 'CHK':
                ref = t.reference_number
                while ref and ref[-1].isalpha():
                    ref = ref[:-1]
                if ref == ck:
                    dms_txn = t
                    break
        if dms_txn and dms_txn.is_matched:
            jay_in_feb_matched.append(c)
        else:
            jay_in_feb_unmatched.append(c)

    matched_total = sum(c['amount'] for c in jay_in_feb_matched)
    unmatched_total = sum(c['amount'] for c in jay_in_feb_unmatched)
    print(f"\n    Of those in Feb DMS:")
    print(f"      Matched by engine:   {len(jay_in_feb_matched)} checks  {fmt(matched_total)}")
    print(f"      Unmatched by engine: {len(jay_in_feb_unmatched)} checks  {fmt(unmatched_total)}")

    if jay_in_feb_matched:
        print(f"\n    WARNING: {len(jay_in_feb_matched)} of Jay's 'outstanding' checks were matched")
        print(f"    by our engine (they cleared the bank). This is expected -- Jay's list")
        print(f"    was built BEFORE these checks cleared.")
        for c in sorted(jay_in_feb_matched, key=lambda x: abs(x['amount']), reverse=True)[:10]:
            print(f"      Check #{c['check_number']}  {fmt(c['amount']):>12s}  {c['payee'][:30]}")

    # Our expected outstanding = Feb DMS checks that are unmatched (didn't clear bank)
    # PLUS all prior-period checks still outstanding
    # Use absolute values for a like-for-like comparison
    our_estimated_os_abs = our_os_check_abs + jay_not_in_feb_total
    print(f"\n  Estimated total outstanding checks from our perspective:")
    print(f"    Our Feb unmatched (abs):  {fmt(our_os_check_abs)}")
    print(f"    Prior period (Jay's):     {fmt(jay_not_in_feb_total)}")
    print(f"    Estimated total:          {fmt(our_estimated_os_abs)}")
    print(f"    Jay's total:              {fmt(jay_os_total)}")
    variance = our_estimated_os_abs - jay_os_total
    print(f"    Variance:                 {fmt(variance)}")

    # Explain the variance
    if abs(variance) > 1.0:
        print(f"\n    Variance explanation:")
        print(f"      Jay's list includes {len(jay_in_feb_matched)} checks that our engine matched")
        print(f"      (cleared the bank in Feb). Those total {fmt(matched_total)}.")
        print(f"      After removing those from Jay's total:")
        adjusted_jay = jay_os_total - matched_total
        print(f"      Jay's adjusted: {fmt(adjusted_jay)}")
        print(f"      Our estimated:  {fmt(our_estimated_os_abs)}")
        print(f"      Residual:       {fmt(our_estimated_os_abs - adjusted_jay)}")
        print(f"      (Residual may reflect DMS self-canceling pairs, void/reissues,")
        print(f"       or checks in Jan GL that carried forward.)")

    # ======================================================================
    # TASK 6: VERIFY SRC=9 BALANCE FORWARD ENTRY
    # ======================================================================
    print("\n" + "=" * 80)
    print("  TASK 6: SRC=9 BALANCE FORWARD VERIFICATION")
    print("=" * 80)

    # Find all SRC=9 entries
    src9 = [t for t in feb_dms if t.type_code == '9']
    balfwd_entries = [t for t in feb_dms if 'BALFWD' in t.reference_number.upper()]

    # Also check among all types
    all_balfwd = [t for t in feb_dms
                  if 'BALFWD' in t.reference_number.upper()
                  or 'BALFWD' in t.description.upper()
                  or t.type_code == '9']

    print(f"\n  SRC=9 entries in Feb GL: {len(src9)}")
    print(f"  BALFWD entries in Feb GL: {len(balfwd_entries)}")
    print(f"  Combined: {len(all_balfwd)}")

    for t in all_balfwd:
        is_matched = t.is_matched
        match_info = ""
        if is_matched:
            # Find the match
            for m in results['all_matches']:
                if t.transaction_id in m.dms_transaction_ids:
                    match_info = (f"MATCHED (ID={m.match_id}, conf={m.confidence_score}%, "
                                  f"type={m.match_type}, breakdown={m.score_breakdown[:60]})")
                    break
        else:
            match_info = "UNMATCHED"

        print(f"\n    ID={t.transaction_id}")
        print(f"    SRC/Type: {t.type_code}")
        print(f"    Ref: {t.reference_number}")
        print(f"    Date: {t.transaction_date}")
        print(f"    Amount: {fmt(t.amount)}")
        print(f"    Description: {t.description}")
        print(f"    Status: {match_info}")

    # Check if it was excluded by Phase -1 (self-canceling)
    from matching_engine import _detect_self_canceling_pairs
    excluded = _detect_self_canceling_pairs(feb_dms)
    for t in all_balfwd:
        if t.transaction_id in excluded:
            print(f"\n    NOTE: ID={t.transaction_id} was EXCLUDED by Phase -1 (self-canceling pair)")

    # ======================================================================
    # SUMMARY: FULL UNMATCHED INVENTORY
    # ======================================================================
    print("\n" + "=" * 80)
    print("  SUMMARY: FULL UNMATCHED INVENTORY")
    print("=" * 80)

    # Unmatched bank by type
    print(f"\n  UNMATCHED BANK ({len(unmatched_bank)} items):")
    ub_by_type = defaultdict(lambda: {'items': [], 'total': 0.0})
    for t in unmatched_bank:
        ub_by_type[t.type_code]['items'].append(t)
        ub_by_type[t.type_code]['total'] += t.amount

    for tc in sorted(ub_by_type.keys()):
        d = ub_by_type[tc]
        print(f"    {tc:10s}: {len(d['items']):3d} items  {fmt(d['total']):>15s}")
        for t in sorted(d['items'], key=lambda x: abs(x.amount), reverse=True)[:3]:
            print(f"      ID={t.transaction_id} {t.transaction_date} {fmt(t.amount):>12s} {t.description[:45]}")
        if len(d['items']) > 3:
            print(f"      ... and {len(d['items']) - 3} more")

    ub_total = sum(t.amount for t in unmatched_bank)
    print(f"    {'TOTAL':10s}: {len(unmatched_bank):3d} items  {fmt(ub_total):>15s}")

    # Unmatched DMS by type
    print(f"\n  UNMATCHED DMS ({len(unmatched_dms)} items):")
    ud_by_type = defaultdict(lambda: {'items': [], 'total': 0.0})
    for t in unmatched_dms:
        ud_by_type[t.type_code]['items'].append(t)
        ud_by_type[t.type_code]['total'] += t.amount

    for tc in sorted(ud_by_type.keys()):
        d = ud_by_type[tc]
        print(f"    {tc:10s}: {len(d['items']):3d} items  {fmt(d['total']):>15s}")
        for t in sorted(d['items'], key=lambda x: abs(x.amount), reverse=True)[:3]:
            print(f"      ID={t.transaction_id} {t.transaction_date} Ref={t.reference_number:12s} "
                  f"{fmt(t.amount):>12s} {t.description[:40]}")
        if len(d['items']) > 3:
            print(f"      ... and {len(d['items']) - 3} more")

    ud_total = sum(t.amount for t in unmatched_dms)
    print(f"    {'TOTAL':10s}: {len(unmatched_dms):3d} items  {fmt(ud_total):>15s}")

    # ======================================================================
    # RECONCILIATION PROOF
    # ======================================================================
    print("\n" + "=" * 80)
    print("  RECONCILIATION PROOF")
    print("=" * 80)

    matched_bank_total = sum(t.amount for t in bank_txns if t.is_matched)
    matched_dms_total = sum(t.amount for t in feb_dms if t.is_matched)
    matched_jan_total = sum(t.amount for t in jan_dms if t.is_matched)

    print(f"\n  Matched bank activity:       {fmt(matched_bank_total)}")
    print(f"  Matched Feb DMS activity:    {fmt(matched_dms_total)}")
    print(f"  Matched Jan DMS activity:    {fmt(matched_jan_total)}")
    print(f"  Match difference (bank-DMS): {fmt(matched_bank_total - matched_dms_total - matched_jan_total)}")

    print(f"\n  Unmatched bank total:        {fmt(ub_total)}")
    print(f"  Unmatched DMS total:         {fmt(ud_total)}")

    # Amount difference across all matches
    total_amt_diff = sum(m.amount_difference for m in results['all_matches'])
    print(f"\n  Sum of amount differences in all matches: {fmt(total_amt_diff)}")

    # Confidence distribution
    high = [m for m in results['all_matches'] if m.confidence_score >= 85]
    med = [m for m in results['all_matches'] if 60 <= m.confidence_score < 85]
    low = [m for m in results['all_matches'] if m.confidence_score < 60]
    print(f"\n  Confidence distribution:")
    print(f"    High (85%+):   {len(high)} ({len(high)/total_matches*100:.1f}%)")
    print(f"    Medium (60-84): {len(med)} ({len(med)/total_matches*100:.1f}%)")
    print(f"    Low (<60):     {len(low)} ({len(low)/total_matches*100:.1f}%)")

    # Auto-accept rate
    auto_accept = len(high) / len(bank_txns) * 100
    bank_match = (len(bank_txns) - len(unmatched_bank)) / len(bank_txns) * 100
    print(f"\n  Auto-accept rate (85%+ / total bank):  {auto_accept:.1f}%")
    print(f"  Bank match rate:                        {bank_match:.1f}%")

    # ======================================================================
    # ENGINE PERFORMANCE ASSESSMENT
    # ======================================================================
    print("\n" + "=" * 80)
    print("  ENGINE PERFORMANCE ASSESSMENT")
    print("=" * 80)

    print(f"""
  The ABR engine processed {len(bank_txns)} bank transactions against
  {len(feb_dms)} Feb GL + {len(jan_dms)} Jan GL entries.

  Results:
    - {total_matches} total matches ({bank_match:.1f}% bank match rate)
    - {len(high)} auto-accept quality ({auto_accept:.1f}% auto-accept rate)
    - {len(unmatched_bank)} unmatched bank items
    - {len(unmatched_dms)} unmatched DMS items

  Unmatched bank items are primarily:
    - Sweep/securities transfers (reconciling items between BofA and Truist)
    - ACH items needing FINDEP matching
    - Checks from prior periods not in Jan/Feb GL

  Unmatched DMS items are primarily:
    - Outstanding checks (didn't clear bank in Feb)
    - FINDEP entries for journal entries and Fulton Bank activity
    - Balance forward entry (SRC=9)

  Jay's $0.00 difference is achievable when:
    1. Outstanding checks + deposits are carried forward correctly
    2. Sweep/securities/Fulton entries are mapped to their bank-side counterparts
    3. FINDEP journal entries are identified and excluded from matching
    """)

    print("=" * 80)
    print("  END OF ACCOUNTING ANALYSIS")
    print("=" * 80)

    return results


if __name__ == '__main__':
    main()
