"""
ABR Matching Engine — Python Reference Implementation

This module implements the exact same matching algorithm that the VBA code uses.
It serves as the authoritative reference for testing and validation.
"""

from __future__ import annotations

import csv
import re
import time
from dataclasses import dataclass, field
from datetime import datetime, date, timedelta
from itertools import combinations
from typing import Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class Transaction:
    transaction_id: int
    source: str  # "BANK" or "DMS"
    transaction_date: date
    description: str
    amount: float
    check_number: str = ""
    reference_number: str = ""
    type_code: str = ""
    post_date: Optional[date] = None
    bank_source: str = ""  # "BOFA" or "TRUIST"
    gl_account: str = ""
    is_matched: bool = False
    match_id: int = 0
    is_reconciling_item: bool = False
    reconciling_type: str = ""  # "SWEEP", "SECURITIES", or ""


@dataclass
class MatchResult:
    match_id: int
    bank_transaction_ids: list[int]
    dms_transaction_ids: list[int]
    confidence_score: float
    match_type: str  # "ONE_TO_ONE", "MANY_TO_ONE_BANK", "MANY_TO_ONE_DMS"
    status: str = "STAGED"  # "STAGED", "ACCEPTED", "REJECTED"
    score_breakdown: str = ""
    amount_difference: float = 0.0
    date_difference: int = 0
    check_number_match: Optional[bool] = None
    bank_amount: float = 0.0
    dms_amount: float = 0.0


@dataclass
class MatchConfig:
    high_confidence_threshold: float = 85.0
    medium_confidence_threshold: float = 60.0
    low_confidence_threshold: float = 40.0
    amount_weight: float = 0.40
    check_number_weight: float = 0.25
    date_proximity_weight: float = 0.25
    description_weight: float = 0.10
    cvr_tolerance: float = 0.01
    date_window_days: int = 7
    max_cvr_fragments: int = 6
    max_cvr_candidates: int = 20
    cvr_timeout_seconds: float = 2.0


# ---------------------------------------------------------------------------
# Scoring functions
# ---------------------------------------------------------------------------

def score_amount(bank_amount: float, dms_amount: float) -> float:
    """Score the amount match factor (0-100). Acts as a gate — >$0.05 diff = 0."""
    diff = abs(bank_amount - dms_amount)
    if diff == 0:
        return 100.0
    elif diff <= 0.01:
        return 98.0
    elif diff <= 0.05:
        return 90.0
    else:
        return 0.0


def score_check_number(bank_check: str, dms_check: str) -> tuple[float, bool]:
    """
    Score check number match (0-100).
    Returns (score, is_veto) where is_veto=True means mismatched check numbers
    should cap the total confidence at 30.
    """
    bank_clean = bank_check.strip()
    dms_clean = dms_check.strip()

    if bank_clean and dms_clean:
        if bank_clean == dms_clean:
            return 100.0, False
        else:
            return 0.0, True  # Veto — mismatched check numbers
    else:
        # One or both missing — inconclusive, not penalized
        return 50.0, False


def score_date(bank_date: date, dms_date: date, max_window: int = 7) -> float:
    """Score date proximity (0-100). Beyond max_window = 0."""
    days_diff = abs((bank_date - dms_date).days)

    if days_diff > max_window:
        return 0.0

    # Scoring curve
    scores = {0: 100, 1: 95, 2: 85, 3: 70, 4: 55, 5: 40, 6: 25, 7: 10}
    return scores.get(days_diff, 0.0)


def levenshtein_distance(s1: str, s2: str) -> int:
    """Compute Levenshtein edit distance between two strings."""
    if len(s1) < len(s2):
        return levenshtein_distance(s2, s1)

    if len(s2) == 0:
        return len(s1)

    prev_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        curr_row = [i + 1]
        for j, c2 in enumerate(s2):
            # Cost is 0 if characters match, 1 otherwise
            insertions = prev_row[j + 1] + 1
            deletions = curr_row[j] + 1
            substitutions = prev_row[j] + (c1 != c2)
            curr_row.append(min(insertions, deletions, substitutions))
        prev_row = curr_row

    return prev_row[-1]


def clean_description(desc: str) -> str:
    """Normalize a description for comparison."""
    cleaned = desc.upper().strip()
    cleaned = re.sub(r'\s+', ' ', cleaned)
    # Remove common noise words
    for noise in ['THE', 'A', 'AN', 'FOR', 'OF', 'TO', 'IN', 'ON', 'AT']:
        cleaned = re.sub(rf'\b{noise}\b', '', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned


def score_description(bank_desc: str, dms_desc: str) -> float:
    """Score description similarity (0-100). Low-weighted tiebreaker."""
    clean_bank = clean_description(bank_desc)
    clean_dms = clean_description(dms_desc)

    if not clean_bank or not clean_dms:
        return 50.0  # Can't compare, neutral

    max_len = max(len(clean_bank), len(clean_dms))
    if max_len == 0:
        return 50.0

    distance = levenshtein_distance(clean_bank, clean_dms)
    similarity = (1 - (distance / max_len)) * 100

    # Bonus for shared significant words
    bank_words = set(clean_bank.split())
    dms_words = set(clean_dms.split())
    shared = bank_words & dms_words
    # Filter out short words (likely noise)
    significant_shared = [w for w in shared if len(w) >= 4]

    if len(significant_shared) >= 2:
        similarity += 20
    elif len(significant_shared) == 1:
        similarity += 10

    # Check/CHK keyword bonus
    check_terms = {'CHECK', 'CHK'}
    if (bank_words & check_terms) and (dms_words & check_terms):
        similarity += 10

    return min(similarity, 100.0)


def score_match(bank_txn: Transaction, dms_txn: Transaction,
                config: MatchConfig) -> Optional[MatchResult]:
    """
    Compute confidence score for a bank-DMS transaction pair.
    Returns None if not a viable match candidate (amount gate fails).
    """
    # Amount gate
    amount_score = score_amount(bank_txn.amount, dms_txn.amount)
    if amount_score == 0:
        return None

    # Check number
    bank_check = bank_txn.check_number
    dms_check = dms_txn.reference_number if dms_txn.type_code == "CHK" else dms_txn.check_number
    check_score, is_veto = score_check_number(bank_check, dms_check)

    # Date
    date_score = score_date(bank_txn.transaction_date, dms_txn.transaction_date,
                            config.date_window_days)

    # Description
    desc_score = score_description(bank_txn.description, dms_txn.description)

    # Weighted composite
    total = (amount_score * config.amount_weight +
             check_score * config.check_number_weight +
             date_score * config.date_proximity_weight +
             desc_score * config.description_weight)

    # Check number veto — cap at 30 if check numbers mismatch
    if is_veto:
        total = min(total, 30.0)

    # Build score breakdown string
    breakdown = (f"Amount:{amount_score:.0f}*{config.amount_weight:.2f} "
                 f"Check:{check_score:.0f}*{config.check_number_weight:.2f} "
                 f"Date:{date_score:.0f}*{config.date_proximity_weight:.2f} "
                 f"Desc:{desc_score:.0f}*{config.description_weight:.2f}")
    if is_veto:
        breakdown += " [CHECK# VETO: capped at 30]"

    days_diff = abs((bank_txn.transaction_date - dms_txn.transaction_date).days)

    return MatchResult(
        match_id=0,  # Assigned later
        bank_transaction_ids=[bank_txn.transaction_id],
        dms_transaction_ids=[dms_txn.transaction_id],
        confidence_score=round(total, 2),
        match_type="ONE_TO_ONE",
        score_breakdown=breakdown,
        amount_difference=round(bank_txn.amount - dms_txn.amount, 2),
        date_difference=days_diff,
        check_number_match=not is_veto if (bank_check and dms_check) else None,
        bank_amount=bank_txn.amount,
        dms_amount=dms_txn.amount,
    )


# ---------------------------------------------------------------------------
# Pass-Rule Matching (v2 architecture)
# ---------------------------------------------------------------------------
# Phase -1: DMS self-canceling pair detection (voids/reversals)
# Phase  0, Rule 0: Check# confirmed + exact amount -> 100%
# Phase  0, Rule 1: Unique exact amount + date corroboration -> 95%
# Phase  0, Rule 1b: Unique exact amount, no date corroboration -> 85%
# Phase  1: Scored matching for duplicate-amount groups (margin-based)
# Phase  4: Near-amount matching ($0.01 tolerance, 55%, always review)
# ---------------------------------------------------------------------------

def _detect_self_canceling_pairs(dms_txns: list[Transaction]) -> set[int]:
    """Phase -1: Identify DMS self-canceling pairs (voids/reversals)."""
    excluded = set()
    for i, a in enumerate(dms_txns):
        if a.transaction_id in excluded or a.is_matched:
            continue
        for j in range(i + 1, len(dms_txns)):
            b = dms_txns[j]
            if b.transaction_id in excluded or b.is_matched:
                continue
            if abs(a.amount) != abs(b.amount) or a.amount == 0:
                continue
            if (a.amount > 0) == (b.amount > 0):
                continue
            # Same reference
            same_ref = False
            if a.check_number and a.check_number == b.check_number:
                same_ref = True
            elif a.reference_number and a.reference_number == b.reference_number:
                same_ref = True
            if same_ref and abs((a.transaction_date - b.transaction_date).days) <= 30:
                excluded.add(a.transaction_id)
                excluded.add(b.transaction_id)
                break
    return excluded


def _get_dms_check(dms: Transaction) -> str:
    """Get the check number for a DMS transaction (reference for CHK type)."""
    if dms.check_number:
        return dms.check_number
    if dms.type_code == 'CHK' and dms.reference_number:
        # Strip trailing letters from reference (e.g., "123095A" -> "123095")
        ref = dms.reference_number
        while ref and ref[-1].isalpha():
            ref = ref[:-1]
        return ref
    return ''


def run_matching(bank_txns: list[Transaction], dms_txns: list[Transaction],
                 config: MatchConfig) -> tuple[list[MatchResult], list[Transaction], list[Transaction]]:
    """
    Run the pass-rule matching pipeline (v2 architecture).

    Phase -1: Exclude DMS self-canceling pairs
    Phase  0: Deterministic pass rules (check#, unique amount)
    Phase  1: Scored matching for duplicate-amount groups
    Phase  4: Near-amount matching ($0.01 tolerance)

    Returns:
        (staged_matches, unmatched_bank, unmatched_dms)
    """
    staged: list[MatchResult] = []
    matched_bank_ids: set[int] = set()
    matched_dms_ids: set[int] = set()
    match_id_counter = 1

    def is_bank_available(t: Transaction) -> bool:
        return not t.is_matched and t.transaction_id not in matched_bank_ids

    def is_dms_available(t: Transaction) -> bool:
        return (not t.is_matched and t.transaction_id not in matched_dms_ids
                and t.transaction_id not in excluded_dms_ids)

    def stage(result: MatchResult):
        nonlocal match_id_counter
        result.match_id = match_id_counter
        match_id_counter += 1
        staged.append(result)
        for bid in result.bank_transaction_ids:
            matched_bank_ids.add(bid)
        for did in result.dms_transaction_ids:
            matched_dms_ids.add(did)

    # Phase -1: Self-canceling pairs
    excluded_dms_ids = _detect_self_canceling_pairs(dms_txns)

    # Phase 0, Rule 0: Check# confirmed + exact amount -> 100%
    for bank in bank_txns:
        if not is_bank_available(bank) or not bank.check_number:
            continue
        for dms in dms_txns:
            if not is_dms_available(dms):
                continue
            dms_check = _get_dms_check(dms)
            if not dms_check:
                continue
            if bank.check_number != dms_check:
                continue
            if bank.amount != dms.amount:
                continue
            # MATCH — 100% confidence
            days_diff = abs((bank.transaction_date - dms.transaction_date).days)
            stage(MatchResult(
                match_id=0,
                bank_transaction_ids=[bank.transaction_id],
                dms_transaction_ids=[dms.transaction_id],
                confidence_score=100.0,
                match_type="ONE_TO_ONE",
                score_breakdown=f"PASS RULE: Check# {bank.check_number} confirmed + "
                                f"exact amount ${bank.amount:,.2f} | {days_diff}d gap -> 100%",
                amount_difference=0.0,
                date_difference=days_diff,
                check_number_match=True,
                bank_amount=bank.amount,
                dms_amount=dms.amount,
            ))
            break  # This bank txn is matched

    # Phase 0, Rule 1: Unique exact amount
    for bank in bank_txns:
        if not is_bank_available(bank):
            continue
        # Count DMS candidates at exact amount
        candidates = [d for d in dms_txns
                      if is_dms_available(d) and d.amount == bank.amount]
        if len(candidates) != 1:
            continue
        dms = candidates[0]
        # Outstanding deposit guard: BATCH deposits clear in 1-2 days.
        # If DMS deposit date is >3 days after bank date, the DMS entry
        # is likely an outstanding deposit that hasn't hit the bank yet.
        if (dms.type_code == 'BATCH'
                and dms.transaction_date > bank.transaction_date + timedelta(days=3)):
            continue
        days_diff = abs((bank.transaction_date - dms.transaction_date).days)
        if days_diff <= 5:
            confidence = 95.0
            rule_tag = f"date match {days_diff}d -> 95%"
        else:
            confidence = 85.0
            rule_tag = f"no date corroboration {days_diff}d -> 85%"

        stage(MatchResult(
            match_id=0,
            bank_transaction_ids=[bank.transaction_id],
            dms_transaction_ids=[dms.transaction_id],
            confidence_score=confidence,
            match_type="ONE_TO_ONE",
            score_breakdown=f"PASS RULE: Unique amount ${bank.amount:,.2f} "
                            f"(1 of 1 candidate) + {rule_tag}",
            amount_difference=0.0,
            date_difference=days_diff,
            check_number_match=None,
            bank_amount=bank.amount,
            dms_amount=dms.amount,
        ))

    # Phase 1: Scored matching for duplicate-amount groups
    # Generate all exact-amount candidate pairs for remaining unmatched
    pair_scores: list[tuple[Transaction, Transaction, float]] = []
    for bank in bank_txns:
        if not is_bank_available(bank):
            continue
        for dms in dms_txns:
            if not is_dms_available(dms):
                continue
            if bank.amount != dms.amount:
                continue
            # Outstanding deposit guard (same as Phase 0 Rule 1)
            if (dms.type_code == 'BATCH'
                    and dms.transaction_date > bank.transaction_date + timedelta(days=3)):
                continue
            # Check number veto: if both sides have check numbers and they
            # differ, this is almost certainly a wrong match — skip the pair
            bank_ck = bank.check_number
            dms_ck = _get_dms_check(dms)
            if bank_ck and dms_ck and bank_ck != dms_ck:
                continue
            # Discrimination score: date proximity (70%) + description (30%)
            days_diff = abs((bank.transaction_date - dms.transaction_date).days)
            date_scores = {0: 100, 1: 95, 2: 88, 3: 78, 4: 68, 5: 58}
            if days_diff <= 5:
                ds = date_scores[days_diff]
            elif days_diff <= 7:
                ds = 45
            elif days_diff <= 14:
                ds = 30
            elif days_diff <= 30:
                ds = 15
            else:
                ds = 5
            desc_sim = score_description(bank.description, dms.description)
            pair_score = ds * 0.7 + desc_sim * 0.3
            pair_scores.append((bank, dms, pair_score))

    # Sort by pair score descending
    pair_scores.sort(key=lambda x: x[2], reverse=True)

    # Build lookup: for each bank txn, all candidate pair scores
    from collections import defaultdict
    bank_candidates: dict[int, list[tuple[int, float]]] = defaultdict(list)
    for bank, dms, ps in pair_scores:
        bank_candidates[bank.transaction_id].append((dms.transaction_id, ps))

    # Greedy assignment with margin-based confidence
    for bank, dms, pair_score in pair_scores:
        if bank.transaction_id in matched_bank_ids or dms.transaction_id in matched_dms_ids:
            continue
        # Find runner-up score for this bank txn
        runner_up = -1.0
        for did, ps in bank_candidates[bank.transaction_id]:
            if did != dms.transaction_id and did not in matched_dms_ids:
                runner_up = ps
                break

        if runner_up < 0:
            confidence = 90.0  # No runner-up
        else:
            margin = pair_score - runner_up
            if margin >= 20:
                confidence = 90.0
            elif margin >= 10:
                confidence = 80.0
            elif margin >= 5:
                confidence = 70.0
            else:
                confidence = 60.0

        days_diff = abs((bank.transaction_date - dms.transaction_date).days)
        margin_str = f"{pair_score - runner_up:.0f}" if runner_up >= 0 else "no runner-up"

        stage(MatchResult(
            match_id=0,
            bank_transaction_ids=[bank.transaction_id],
            dms_transaction_ids=[dms.transaction_id],
            confidence_score=confidence,
            match_type="ONE_TO_ONE",
            score_breakdown=f"SCORED: Amount ${bank.amount:,.2f} | "
                            f"Date:{days_diff}d | Margin:{margin_str} -> {confidence:.0f}%",
            amount_difference=0.0,
            date_difference=days_diff,
            check_number_match=None,
            bank_amount=bank.amount,
            dms_amount=dms.amount,
        ))

    # Phase 4: Near-amount matching ($0.01 tolerance)
    for bank in bank_txns:
        if not is_bank_available(bank):
            continue
        best_dms = None
        best_diff = 999.0
        for dms in dms_txns:
            if not is_dms_available(dms):
                continue
            diff = abs(bank.amount - dms.amount)
            if 0 < diff <= 0.01 and diff < best_diff:
                best_dms = dms
                best_diff = diff
        if best_dms:
            days_diff = abs((bank.transaction_date - best_dms.transaction_date).days)
            stage(MatchResult(
                match_id=0,
                bank_transaction_ids=[bank.transaction_id],
                dms_transaction_ids=[best_dms.transaction_id],
                confidence_score=55.0,
                match_type="ONE_TO_ONE",
                score_breakdown=f"NEAR AMOUNT: ${bank.amount:,.2f} vs "
                                f"${best_dms.amount:,.2f} (diff=${best_diff:.2f}) | "
                                f"{days_diff}d -> 55%",
                amount_difference=round(bank.amount - best_dms.amount, 2),
                date_difference=days_diff,
                check_number_match=None,
                bank_amount=bank.amount,
                dms_amount=best_dms.amount,
            ))

    # Mark matched transactions
    for bank in bank_txns:
        if bank.transaction_id in matched_bank_ids:
            bank.is_matched = True
            for m in staged:
                if bank.transaction_id in m.bank_transaction_ids:
                    bank.match_id = m.match_id
                    break

    for dms in dms_txns:
        if dms.transaction_id in matched_dms_ids:
            dms.is_matched = True
            for m in staged:
                if dms.transaction_id in m.dms_transaction_ids:
                    dms.match_id = m.match_id
                    break

    unmatched_bank = [b for b in bank_txns if not b.is_matched]
    unmatched_dms = [d for d in dms_txns if not d.is_matched]

    return staged, unmatched_bank, unmatched_dms


# ---------------------------------------------------------------------------
# CVR / Many-to-One Matching
# ---------------------------------------------------------------------------

def find_subset_sum(candidates: list[Transaction], target: float,
                    tolerance: float = 0.01, max_depth: int = 6,
                    timeout: float = 2.0) -> list[list[Transaction]]:
    """
    Find all subsets of candidates whose amounts sum to target within tolerance.

    Uses iterative deepening to find combinations of 2..max_depth items.
    Respects timeout to prevent Excel freezing.

    Returns list of valid subsets (each subset is a list of Transactions).
    """
    results: list[list[Transaction]] = []
    start_time = time.time()

    # Sort by amount descending for better pruning
    sorted_candidates = sorted(candidates, key=lambda t: abs(t.amount), reverse=True)

    for depth in range(2, max_depth + 1):
        if time.time() - start_time > timeout:
            break

        for combo in combinations(sorted_candidates, depth):
            if time.time() - start_time > timeout:
                break

            combo_sum = sum(t.amount for t in combo)
            if abs(combo_sum - target) <= tolerance:
                results.append(list(combo))

    return results


def score_cvr_group(group: list[Transaction], target_txn: Transaction,
                    tolerance: float = 0.01) -> float:
    """
    Compute confidence score for a CVR group match.
    Weighted: SumAccuracy (50%) + DateClustering (30%) + FragmentCount (20%)
    """
    group_sum = sum(t.amount for t in group)
    variance = abs(group_sum - target_txn.amount)

    # Sum accuracy score (0-100)
    if variance == 0:
        sum_score = 100.0
    elif variance <= 0.01:
        sum_score = 95.0
    elif variance <= tolerance:
        sum_score = 80.0
    else:
        sum_score = 0.0

    # Date clustering score (0-100)
    if group:
        dates = [t.transaction_date for t in group]
        date_spread = (max(dates) - min(dates)).days
        if date_spread <= 1:
            date_score = 100.0
        elif date_spread <= 3:
            date_score = 80.0
        elif date_spread <= 5:
            date_score = 60.0
        else:
            date_score = 30.0
    else:
        date_score = 0.0

    # Fragment count score (0-100)
    n = len(group)
    fragment_scores = {2: 100, 3: 85, 4: 65, 5: 45, 6: 45}
    frag_score = fragment_scores.get(n, 30.0)

    # Weighted composite
    total = sum_score * 0.50 + date_score * 0.30 + frag_score * 0.20
    return round(total, 2)


def run_cvr_matching(unmatched_bank: list[Transaction],
                     unmatched_dms: list[Transaction],
                     config: MatchConfig,
                     next_match_id: int = 1000) -> list[MatchResult]:
    """
    Run CVR many-to-one matching: find groups of bank transactions
    that sum to a single DMS transaction.

    Only considers DMS transactions with type_code 'CVR' or amount > $5000.
    """
    cvr_matches: list[MatchResult] = []
    match_id = next_match_id

    # Identify CVR candidates on the DMS side
    cvr_dms = [d for d in unmatched_dms
               if d.type_code == "CVR" or abs(d.amount) > 5000]

    for dms_txn in cvr_dms:
        if dms_txn.is_matched:
            continue

        # Find bank candidates: unmatched, same sign, smaller amount,
        # within date window
        bank_candidates = []
        for bank in unmatched_bank:
            if bank.is_matched:
                continue
            # Same sign (both positive or both negative)
            if (bank.amount > 0) != (dms_txn.amount > 0):
                continue
            # Fragment must be smaller than the whole
            if abs(bank.amount) >= abs(dms_txn.amount):
                continue
            # Within date window
            days_diff = abs((bank.transaction_date - dms_txn.transaction_date).days)
            if days_diff <= config.date_window_days:
                bank_candidates.append(bank)

        if len(bank_candidates) < 2:
            continue

        # Limit candidates
        bank_candidates = bank_candidates[:config.max_cvr_candidates]

        # Find subsets that sum to DMS amount
        subsets = find_subset_sum(
            bank_candidates,
            dms_txn.amount,
            tolerance=config.cvr_tolerance,
            max_depth=config.max_cvr_fragments,
            timeout=config.cvr_timeout_seconds
        )

        for subset in subsets:
            confidence = score_cvr_group(subset, dms_txn, config.cvr_tolerance)

            if confidence < config.low_confidence_threshold:
                continue

            group_sum = sum(t.amount for t in subset)
            bank_ids = [t.transaction_id for t in subset]

            result = MatchResult(
                match_id=match_id,
                bank_transaction_ids=bank_ids,
                dms_transaction_ids=[dms_txn.transaction_id],
                confidence_score=confidence,
                match_type="MANY_TO_ONE_BANK",
                score_breakdown=f"CVR group: {len(subset)} fragments, "
                                f"sum=${group_sum:.2f} vs target=${dms_txn.amount:.2f}, "
                                f"variance=${abs(group_sum - dms_txn.amount):.2f}",
                amount_difference=round(group_sum - dms_txn.amount, 2),
                bank_amount=group_sum,
                dms_amount=dms_txn.amount,
            )
            cvr_matches.append(result)
            match_id += 1

    return cvr_matches


def run_reverse_split_matching(unmatched_bank: list[Transaction],
                               unmatched_dms: list[Transaction],
                               config: MatchConfig,
                               next_match_id: int = 2000) -> list[MatchResult]:
    """
    Run reverse split matching: find groups of DMS transactions
    that sum to a single bank transaction.
    """
    split_matches: list[MatchResult] = []
    match_id = next_match_id

    # Large bank deposits that have no match
    large_bank = [b for b in unmatched_bank
                  if not b.is_matched and abs(b.amount) > 5000]

    for bank_txn in large_bank:
        dms_candidates = []
        for dms in unmatched_dms:
            if dms.is_matched:
                continue
            if (dms.amount > 0) != (bank_txn.amount > 0):
                continue
            if abs(dms.amount) >= abs(bank_txn.amount):
                continue
            days_diff = abs((dms.transaction_date - bank_txn.transaction_date).days)
            if days_diff <= config.date_window_days:
                dms_candidates.append(dms)

        if len(dms_candidates) < 2:
            continue

        dms_candidates = dms_candidates[:config.max_cvr_candidates]

        subsets = find_subset_sum(
            dms_candidates,
            bank_txn.amount,
            tolerance=config.cvr_tolerance,
            max_depth=config.max_cvr_fragments,
            timeout=config.cvr_timeout_seconds
        )

        for subset in subsets:
            confidence = score_cvr_group(subset, bank_txn, config.cvr_tolerance)

            if confidence < config.low_confidence_threshold:
                continue

            group_sum = sum(t.amount for t in subset)
            dms_ids = [t.transaction_id for t in subset]

            result = MatchResult(
                match_id=match_id,
                bank_transaction_ids=[bank_txn.transaction_id],
                dms_transaction_ids=dms_ids,
                confidence_score=confidence,
                match_type="MANY_TO_ONE_DMS",
                score_breakdown=f"Reverse split: {len(subset)} DMS entries, "
                                f"sum=${group_sum:.2f} vs bank=${bank_txn.amount:.2f}",
                amount_difference=round(group_sum - bank_txn.amount, 2),
                bank_amount=bank_txn.amount,
                dms_amount=group_sum,
            )
            split_matches.append(result)
            match_id += 1

    return split_matches


# ---------------------------------------------------------------------------
# CSV / XLSX Parsers — Real-World Formats
# ---------------------------------------------------------------------------

def extract_check_number(description: str) -> str:
    """Extract check number from a bank description field."""
    patterns = [
        r'CHECK\s*#?\s*(\d{3,8})',
        r'CHK\s*#?\s*(\d{3,8})',
        r'CHECK\s+(\d{3,8})',
        r'CK\s*#?\s*(\d{3,8})',
    ]
    desc_upper = description.upper()
    for pattern in patterns:
        match = re.search(pattern, desc_upper)
        if match:
            return match.group(1)
    return ""


def _parse_amount(val: str) -> Optional[float]:
    """Parse a dollar amount string, handling commas, quotes, parens, dollar signs."""
    if not val or not val.strip():
        return None
    cleaned = val.strip().replace('"', '').replace('$', '').replace(',', '').replace(' ', '')
    if cleaned.startswith('(') and cleaned.endswith(')'):
        cleaned = '-' + cleaned[1:-1]
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_check_date(date_str: str, statement_year: int = 2025) -> Optional[date]:
    """Parse dates in D-Mon format (e.g., '16-May', '2-May') used in BofA Checks section."""
    date_str = date_str.strip().replace('*', '')
    # Try D-Mon format (Checks section)
    try:
        parsed = datetime.strptime(date_str, '%d-%b')
        return parsed.replace(year=statement_year).date()
    except ValueError:
        pass
    # Try M/D/YYYY format (Deposits/Withdrawals section)
    try:
        return datetime.strptime(date_str, '%m/%d/%Y').date()
    except ValueError:
        pass
    return None


def parse_bofa_csv(filepath: str) -> list[Transaction]:
    """
    Parse Bank of America COMMERCIAL bank statement CSV.

    Real format is sectioned:
      Row 1: Statement Information header
      Row 2: Account Summary
      Section "Deposits and other credits": Type, Date(M/D/YYYY), DepID, Amount, Description, Ref
      Section "Withdrawals and other Debits": Type, Date(M/D/YYYY), empty, Amount(neg), Description, Ref
      Section "Checks": Type, Date(D-Mon), CheckNum(*), Amount(neg), empty, Ref
      Section "Daily Ledger Balances": skip
    """
    transactions = []
    txn_id = 1

    with open(filepath, 'r', newline='', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) < 5:
                continue

            section = row[0].strip()

            # Skip non-transaction rows
            if section in ('Statement Information', 'Account Summary',
                           'Daily Ledger Balances', ''):
                continue

            date_str = row[1].strip() if len(row) > 1 else ''
            col3 = row[2].strip() if len(row) > 2 else ''
            amount_str = row[3].strip() if len(row) > 3 else ''
            description = row[4].strip() if len(row) > 4 else ''

            # Parse date — Checks use D-Mon, others use M/D/YYYY
            txn_date = _parse_check_date(date_str)
            if txn_date is None:
                continue

            # Parse amount
            amount = _parse_amount(amount_str)
            if amount is None:
                continue

            check_number = ''
            txn_type = ''

            if section == 'Deposits and other credits':
                txn_type = 'DEP'
                # col3 is deposit ID (often "1" for preencoded) — not a check number
            elif section == 'Withdrawals and other Debits':
                txn_type = 'WDR'
                # amount should already be negative
            elif section == 'Checks':
                txn_type = 'CHK'
                # col3 is the check number (may have * suffix)
                check_number = col3.replace('*', '').strip()
                # Checks should be negative
                if amount > 0:
                    amount = -amount
            else:
                continue

            # Extract check number from description if not already found
            if not check_number:
                check_number = extract_check_number(description)

            txn = Transaction(
                transaction_id=txn_id,
                source="BANK",
                transaction_date=txn_date,
                description=description if description else section,
                amount=amount,
                check_number=check_number,
                type_code=txn_type,
                bank_source="BOFA",
            )
            transactions.append(txn)
            txn_id += 1

    return transactions


def parse_truist_csv(filepath: str) -> list[Transaction]:
    """Parse Truist CSV export (Debit/Credit columns)."""
    transactions = []
    txn_id = 1

    with open(filepath, 'r', newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            date_str = row.get('Date', '').strip()
            desc = row.get('Description', '').strip()
            debit_str = row.get('Debit', '').strip().replace(',', '')
            credit_str = row.get('Credit', '').strip().replace(',', '')

            try:
                txn_date = datetime.strptime(date_str, '%m/%d/%Y').date()
            except ValueError:
                try:
                    txn_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    continue

            if debit_str:
                try:
                    amount = -abs(float(debit_str))
                except ValueError:
                    continue
            elif credit_str:
                try:
                    amount = abs(float(credit_str))
                except ValueError:
                    continue
            else:
                continue

            check_num = extract_check_number(desc)

            txn = Transaction(
                transaction_id=txn_id,
                source="BANK",
                transaction_date=txn_date,
                description=desc,
                amount=amount,
                check_number=check_num,
                bank_source="TRUIST",
            )
            transactions.append(txn)
            txn_id += 1

    return transactions


def _excel_serial_to_date(serial: int) -> date:
    """Convert Excel date serial number to Python date."""
    # Excel epoch: 1899-12-30 (with the Lotus 1-2-3 bug for 1900-02-29)
    excel_epoch = date(1899, 12, 30)
    return excel_epoch + timedelta(days=int(serial))


def parse_dms_xlsx(filepath: str) -> list[Transaction]:
    """
    Parse R&R DMS GL export in XLSX format.

    Real format (9 columns):
      SRC | Reference# | Date | Port | Control# | Debit Amount | Credit Amount | Name | Description

    SRC codes: 5=batch, 6=checks/individual, 11=finance deposits
    Debit Amount = positive (money in), Credit Amount = negative (money out)
    """
    if openpyxl is None:
        raise ImportError("openpyxl required for xlsx parsing")

    transactions = []
    txn_id = 1

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        src = str(row[0]).strip()
        reference = str(row[1]).strip() if row[1] else ''
        txn_date_raw = row[2]
        port = str(row[3]).strip() if row[3] else ''
        control = str(row[4]).strip() if row[4] else ''
        debit_amt = row[5]
        credit_amt = row[6]
        name = str(row[7]).strip() if row[7] else ''
        description = str(row[8]).strip() if row[8] else ''

        # Parse date
        if isinstance(txn_date_raw, datetime):
            txn_date = txn_date_raw.date()
        elif isinstance(txn_date_raw, date):
            txn_date = txn_date_raw
        else:
            continue

        # Combine debit/credit into single amount
        # Debit = positive (money received), Credit = negative (money paid out)
        amount = 0.0
        if debit_amt is not None and debit_amt != '' and debit_amt != 0:
            amount = float(debit_amt)
        elif credit_amt is not None and credit_amt != '' and credit_amt != 0:
            amount = float(credit_amt)  # Already negative in the data
        else:
            continue

        # Determine type code from SRC and reference pattern
        type_code = ''
        check_number = ''
        if src == '6':
            # SRC 6 = individual transactions (usually checks)
            type_code = 'CHK'
            # Reference is the check number for SRC=6
            check_match = re.match(r'^(\d{3,8}[A-Z]?)$', reference)
            if check_match:
                # Strip any trailing letter suffix (e.g., "231557A")
                check_number = re.match(r'^(\d+)', reference).group(1)
        elif src == '5':
            # SRC 5 = batch transactions
            if reference.startswith('CA'):
                type_code = 'CASH'
            elif reference.startswith('CK'):
                type_code = 'CKBATCH'
            elif reference.startswith('V'):
                type_code = 'VENDOR'
            else:
                type_code = 'BATCH'
        elif src == '11':
            type_code = 'FINDEP'
        else:
            type_code = src

        # Build description from name + description fields
        full_desc = name
        if description:
            full_desc = f"{name} - {description}" if name else description

        txn = Transaction(
            transaction_id=txn_id,
            source="DMS",
            transaction_date=txn_date,
            description=full_desc,
            amount=amount,
            check_number=check_number,
            reference_number=reference,
            type_code=type_code,
        )
        transactions.append(txn)
        txn_id += 1

    wb.close()
    return transactions


def parse_dms_csv(filepath: str) -> list[Transaction]:
    """Parse R&R DMS GL export — dispatches to xlsx or csv parser."""
    if filepath.endswith('.xlsx') or filepath.endswith('.xls'):
        return parse_dms_xlsx(filepath)

    # Legacy CSV fallback for test data
    transactions = []
    txn_id = 1
    with open(filepath, 'r', newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            date_str = row.get('GL Date', '').strip()
            desc = row.get('Description', '').strip()
            ref = row.get('Reference', '').strip()
            amount_str = row.get('Amount', '0').strip().replace(',', '')
            type_code = row.get('Type Code', '').strip()
            try:
                txn_date = datetime.strptime(date_str, '%m/%d/%Y').date()
            except ValueError:
                try:
                    txn_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    continue
            try:
                amount = float(amount_str)
            except ValueError:
                continue
            check_num = ""
            if type_code == "CHK":
                check_match = re.search(r'(\d{3,8})', ref)
                if check_match:
                    check_num = check_match.group(1)
            txn = Transaction(
                transaction_id=txn_id, source="DMS", transaction_date=txn_date,
                description=desc, amount=amount, check_number=check_num,
                reference_number=ref, type_code=type_code,
            )
            transactions.append(txn)
            txn_id += 1
    return transactions


def parse_outstanding_xlsx(filepath: str) -> list[Transaction]:
    """
    Parse outstanding checks XLSX file.

    Format: Check# | Bank Code | Check Date | Amount | Payee | Cancel Date
    Check Date is Excel serial number. Cancel Date is serial or None.
    Returns only checks that are STILL outstanding (no Cancel Date).
    """
    if openpyxl is None:
        raise ImportError("openpyxl required for xlsx parsing")

    transactions = []
    txn_id = 1

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        check_num = str(int(row[0])) if isinstance(row[0], (int, float)) else str(row[0])
        check_date_raw = row[2]
        amount = row[3]
        payee = str(row[4]) if row[4] else ''
        cancel_date = row[5]

        # Parse check date (Excel serial number)
        if isinstance(check_date_raw, (int, float)):
            txn_date = _excel_serial_to_date(int(check_date_raw))
        elif isinstance(check_date_raw, datetime):
            txn_date = check_date_raw.date()
        elif isinstance(check_date_raw, date):
            txn_date = check_date_raw
        else:
            continue

        if amount is None:
            continue

        # Track whether this check has cleared
        is_cleared = cancel_date is not None

        txn = Transaction(
            transaction_id=txn_id,
            source="OUTSTANDING",
            transaction_date=txn_date,
            description=payee,
            amount=-abs(float(amount)),  # Outstanding checks are outflows (negative)
            check_number=check_num,
            type_code="CLEARED" if is_cleared else "OUTSTANDING",
        )
        transactions.append(txn)
        txn_id += 1

    wb.close()
    return transactions


def parse_outstanding_from_rec(filepath: str,
                               sheet_name: str = "OS CKS") -> list[Transaction]:
    """
    Parse outstanding checks from a Honda-format bank reconciliation XLSX.

    The "OS CKS" sheet has a sparse layout:
      A: Year (sparse — only populated when year changes)
      B: Month (sparse — only populated when month changes)
      C: Check#
      F: Amount (positive in the sheet; we negate to represent outflows)
      H: Payee

    Rows without a check number in column C are blank/separator rows and are
    skipped.  The last row typically contains a total in column F with no
    check number — this is also skipped by the C-is-None guard.

    Returns a list of Transaction objects with:
      source       = "OUTSTANDING"
      type_code    = "OUTSTANDING"
      amount       = negative (outflow)
      check_number = from column C
      description  = payee from column H
      transaction_date = date(1900,1,1) placeholder (the sheet has no date per
                         check; the year/month columns are sparse grouping
                         headers, not per-row dates)
    """
    if openpyxl is None:
        raise ImportError("openpyxl required for xlsx parsing")

    transactions: list[Transaction] = []
    txn_id = 1

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Column C (index 2) is the check number — skip blank rows
        if row[2] is None:
            continue

        check_num = str(int(row[2])) if isinstance(row[2], (int, float)) else str(row[2]).strip()
        if not check_num:
            continue

        # Column F (index 5) is the amount
        amount_raw = row[5] if len(row) > 5 else None
        if amount_raw is None:
            continue

        try:
            amount = -abs(float(amount_raw))  # Outstanding checks are outflows
        except (ValueError, TypeError):
            continue

        # Column H (index 7) is the payee
        payee = ""
        if len(row) > 7 and row[7] is not None:
            payee = str(row[7]).strip()

        txn = Transaction(
            transaction_id=txn_id,
            source="OUTSTANDING",
            transaction_date=date(1900, 1, 1),  # Placeholder — no per-check date
            description=payee,
            amount=amount,
            check_number=check_num,
            type_code="OUTSTANDING",
        )
        transactions.append(txn)
        txn_id += 1

    wb.close()
    return transactions


def parse_bofa_bai_csv(filepath: str) -> list[Transaction]:
    """
    Parse Bank of America BAI flat columnar CSV (e.g., Honda Feb 2026).

    14 columns with header row:
      Account Name, Account Number, Amount(signed), BAI Code, ABA,
      Bank Reference, Currency, Customer Reference, Debit/Credit,
      As of Date(MM/DD/YY), Status, Transaction Description,
      Transaction Detail, Type
    """
    transactions = []
    txn_id = 1

    with open(filepath, 'r', newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Amount: pre-signed (negative = debit)
            amount_str = row.get('Amount', '').strip()
            if not amount_str:
                continue
            try:
                amount = float(amount_str)
            except ValueError:
                continue

            # Date: MM/DD/YY format
            date_str = row.get('As of Date', '').strip()
            try:
                txn_date = datetime.strptime(date_str, '%m/%d/%y').date()
            except ValueError:
                try:
                    txn_date = datetime.strptime(date_str, '%m/%d/%Y').date()
                except ValueError:
                    continue

            # BAI code determines transaction type
            bai_code = row.get('BAI Code', '').strip()

            # Description
            desc = row.get('Transaction Description', '').strip()
            txn_type = row.get('Type', '').strip()
            if txn_type and txn_type.upper() != desc.upper():
                desc = f"{desc} - {txn_type}"

            # Check number: Customer Reference for BAI 475 (checks) only
            check_number = ''
            if bai_code == '475':
                cust_ref = row.get('Customer Reference', '').strip()
                if cust_ref and cust_ref != '0' and cust_ref.isdigit():
                    check_number = cust_ref

            # Derive type code from BAI code
            bai_type_map = {
                '475': 'CHK',       # Checks paid
                '451': 'ACH_DR',    # ACH debit
                '142': 'ACH_CR',    # ACH credit
                '301': 'DEP',       # Commercial deposit
                '501': 'SWEEP',     # Sweep transfer
                '233': 'SEC',       # Securities
                '555': 'NSF',       # Deposited item returned
                '661': 'FEE',       # Account analysis fee
            }
            type_code = bai_type_map.get(bai_code, 'OTHER')

            # Detect reconciling items: sweep transfers (501) and securities (233)
            is_reconciling = bai_code in ('501', '233')
            reconc_type = ""
            if bai_code == '501':
                reconc_type = "SWEEP"
            elif bai_code == '233':
                reconc_type = "SECURITIES"

            txn = Transaction(
                transaction_id=txn_id,
                source="BANK",
                transaction_date=txn_date,
                description=desc,
                amount=amount,
                check_number=check_number,
                type_code=type_code,
                bank_source="BOFA",
                is_reconciling_item=is_reconciling,
                reconciling_type=reconc_type,
            )
            transactions.append(txn)
            txn_id += 1

    return transactions


def detect_bank_format(filepath: str) -> str:
    """Auto-detect bank CSV format by reading the header/first row."""
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        first_line = f.readline().strip().lower()

    if 'statement information' in first_line:
        return "BOFA"
    # BAI flat format — must check before Truist (BAI header contains "debit/credit")
    elif 'bai code' in first_line:
        return "BOFA_BAI"
    elif 'debit' in first_line and 'credit' in first_line:
        return "TRUIST"
    elif 'deposits and other credits' in first_line:
        return "BOFA"
    else:
        return "UNKNOWN"


# ---------------------------------------------------------------------------
# Full reconciliation pipeline
# ---------------------------------------------------------------------------

def run_prior_outstanding_matching(
        unmatched_bank: list[Transaction],
        prior_outstanding_txns: list[Transaction],
        next_match_id: int = 3000,
) -> list[MatchResult]:
    """
    Phase 5: Match remaining unmatched bank checks against a prior month's
    outstanding checks list.

    Logic: A prior-period outstanding check is a GL entry (check was cut in a
    prior month) that had not yet cleared the bank as of the prior month-end.
    If it clears the bank in the current month, it appears on the bank
    statement but NOT in the current or even prior month GL.  The only record
    of it is the prior month's outstanding checks list.

    Match criteria:
      - Bank transaction must be a check (has check_number)
      - Check number must match exactly
      - Amount must match exactly (both should be negative)
      - 100% confidence (check# confirmed from outstanding list)

    The prior_outstanding_txns should come from the PRIOR month's bank rec
    outstanding checks list — i.e. checks that were outstanding at the end of
    last month.  If a check was outstanding last month and now appears on the
    bank statement this month, it cleared this month.

    Returns list of MatchResult with match_type="PRIOR_OUTSTANDING".
    """
    matches: list[MatchResult] = []
    match_id = next_match_id

    # Index outstanding checks by check number for fast lookup
    os_by_check: dict[str, list[Transaction]] = {}
    for os_txn in prior_outstanding_txns:
        os_by_check.setdefault(os_txn.check_number, []).append(os_txn)

    matched_os_ids: set[int] = set()

    for bank in unmatched_bank:
        if bank.is_matched or not bank.check_number:
            continue

        candidates = os_by_check.get(bank.check_number, [])
        for os_txn in candidates:
            if os_txn.transaction_id in matched_os_ids:
                continue
            # Amount must match exactly (both negative)
            if bank.amount != os_txn.amount:
                continue

            # Match found
            result = MatchResult(
                match_id=match_id,
                bank_transaction_ids=[bank.transaction_id],
                dms_transaction_ids=[os_txn.transaction_id],
                confidence_score=100.0,
                match_type="PRIOR_OUTSTANDING",
                score_breakdown=(
                    f"PRIOR OUTSTANDING: Check# {bank.check_number} confirmed "
                    f"from prior month outstanding list + exact amount "
                    f"${bank.amount:,.2f} -> 100%"
                ),
                amount_difference=0.0,
                date_difference=0,
                check_number_match=True,
                bank_amount=bank.amount,
                dms_amount=os_txn.amount,
            )
            matches.append(result)
            match_id += 1
            bank.is_matched = True
            bank.match_id = result.match_id
            matched_os_ids.add(os_txn.transaction_id)
            break  # Bank check matched

    return matches


def run_full_reconciliation(bank_txns: list[Transaction],
                            dms_txns: list[Transaction],
                            config: MatchConfig | None = None,
                            prior_dms_txns: list[Transaction] | None = None,
                            prior_outstanding_txns: list[Transaction] | None = None,
                            ) -> dict:
    """
    Run the complete reconciliation pipeline:
    1. 1:1 matching against current month GL
    2. 1:1 matching against prior month GL (fallback for carryover)
    3. CVR many-to-one matching on remaining unmatched
    4. Reverse split matching on remaining unmatched
    5. Prior-period outstanding check matching (if prior outstanding list provided)

    Args:
        bank_txns: Bank statement transactions
        dms_txns: Current month DMS/GL transactions
        config: Matching configuration
        prior_dms_txns: Prior month DMS/GL transactions (optional)
        prior_outstanding_txns: Prior month's outstanding checks list (optional).
            These are checks that were cut in a prior period and had NOT yet
            cleared the bank as of the prior month-end.  If provided, they are
            used in Phase 5 to match bank checks that cleared this month but
            were posted to GL before our available GL data window.
            To obtain this data, ask the controller for the PRIOR month's bank
            reconciliation file and parse its "OS CKS" sheet using
            parse_outstanding_from_rec().

    Returns a dict with all results.
    """
    if config is None:
        config = MatchConfig()

    # Separate reconciling items (sweep transfers, securities) from matching pool
    reconciling_items = [t for t in bank_txns if t.is_reconciling_item]
    matchable_bank_txns = [t for t in bank_txns if not t.is_reconciling_item]

    # Phase 1: 1:1 matching against current month GL
    one_to_one, unmatched_bank, unmatched_dms = run_matching(
        matchable_bank_txns, dms_txns, config
    )

    # Phase 2: Prior-month GL fallback
    prior_matches = []
    if prior_dms_txns:
        prior_one_to_one, unmatched_bank, _ = run_matching(
            unmatched_bank, prior_dms_txns, config
        )
        # Tag as prior-period matches
        for m in prior_one_to_one:
            m.match_type = "PRIOR_PERIOD"
        prior_matches = prior_one_to_one

    # Phase 3: CVR many-to-one (bank fragments → DMS lump sum)
    all_so_far = one_to_one + prior_matches
    next_id = max((m.match_id for m in all_so_far), default=0) + 1000
    cvr_matches = run_cvr_matching(
        unmatched_bank, unmatched_dms, config,
        next_match_id=next_id
    )

    # Phase 4: Reverse split (DMS fragments → bank lump sum)
    next_id = max((m.match_id for m in cvr_matches), default=next_id) + 1000
    split_matches = run_reverse_split_matching(
        unmatched_bank, unmatched_dms, config,
        next_match_id=next_id
    )

    # Phase 5: Prior-period outstanding check matching
    prior_outstanding_matches = []
    if prior_outstanding_txns:
        # Recompute unmatched bank after phases 1-4
        all_so_far_ids = set()
        for m in one_to_one + prior_matches + cvr_matches + split_matches:
            all_so_far_ids.update(m.bank_transaction_ids)
        still_unmatched_bank = [b for b in matchable_bank_txns
                                if b.transaction_id not in all_so_far_ids]

        next_id = max((m.match_id for m in split_matches), default=next_id) + 1000
        prior_outstanding_matches = run_prior_outstanding_matching(
            still_unmatched_bank, prior_outstanding_txns,
            next_match_id=next_id,
        )

    all_matches = (one_to_one + prior_matches + cvr_matches +
                   split_matches + prior_outstanding_matches)

    # Final unmatched
    all_matched_bank_ids = set()
    all_matched_dms_ids = set()
    for m in all_matches:
        all_matched_bank_ids.update(m.bank_transaction_ids)
        all_matched_dms_ids.update(m.dms_transaction_ids)

    final_unmatched_bank = [b for b in matchable_bank_txns
                            if b.transaction_id not in all_matched_bank_ids]
    final_unmatched_dms = [d for d in dms_txns
                           if d.transaction_id not in all_matched_dms_ids]

    # Match rate is based on matchable transactions (excluding reconciling items)
    total = len(matchable_bank_txns) + len(dms_txns)
    matched_count = len(all_matched_bank_ids) + len(all_matched_dms_ids)
    match_rate = (matched_count / total * 100) if total > 0 else 0

    # Summarize reconciling items by type
    reconc_by_type: dict[str, list[Transaction]] = {}
    for t in reconciling_items:
        reconc_by_type.setdefault(t.reconciling_type, []).append(t)

    return {
        "one_to_one_matches": one_to_one,
        "prior_period_matches": prior_matches,
        "cvr_matches": cvr_matches,
        "split_matches": split_matches,
        "prior_outstanding_matches": prior_outstanding_matches,
        "all_matches": all_matches,
        "unmatched_bank": final_unmatched_bank,
        "unmatched_dms": final_unmatched_dms,
        "reconciling_items": reconciling_items,
        "reconciling_by_type": reconc_by_type,
        "total_bank": len(bank_txns),  # Total including reconciling
        "total_bank_matchable": len(matchable_bank_txns),  # Excluding reconciling
        "total_dms": len(dms_txns),
        "total_reconciling": len(reconciling_items),
        "match_rate": round(match_rate, 2),
    }
