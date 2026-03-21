"""
Honda Feb 2026 — Full Pipeline Test Against Ground Truth

Runs the ABR matching engine against Honda Feb 2026 real data and
validates results against the completed bank reconciliation (ground truth).

Ground truth from HONDA BANK REC 0226.xlsx:
  - 661 outstanding AP checks ($528,506.53)
  - 12 outstanding payroll checks ($6,950.99)
  - 6 outstanding deposits ($274,213.56)
  - 13 outstanding journal entries (-$108,097.26)
  - Reconciling difference: $0.00
"""

import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

from matching_engine import (
    parse_bofa_bai_csv,
    parse_dms_xlsx,
    parse_outstanding_from_rec,
    detect_bank_format,
    run_full_reconciliation,
    MatchConfig,
    Transaction,
)

# File paths
DOCS = os.path.join(os.path.dirname(__file__), '..', 'docs')
BANK_FILE = os.path.join(DOCS, 'Honda Bkstmt.csv')
FEB_GL_FILE = os.path.join(DOCS, 'honda Feb GL.xlsx')
JAN_GL_FILE = os.path.join(DOCS, 'honda Jan GL.xlsx')
REC_FILE = os.path.join(DOCS, 'HONDA BANK REC 0226.xlsx')

# January 2026 outstanding checks fixture (test data)
FIXTURES = os.path.join(os.path.dirname(__file__), 'fixtures')
JAN_REC_FILE = os.path.join(FIXTURES, 'HONDA_BANK_REC_0126_TEST.xlsx')


def load_outstanding_checks_from_rec():
    """
    Load outstanding check numbers from the bank rec (ground truth).
    Returns set of check number strings.
    """
    try:
        import openpyxl
    except ImportError:
        print("WARNING: openpyxl not available, skipping outstanding checks validation")
        return set(), set(), []

    rec_file = os.path.join(DOCS, 'HONDA BANK REC 0226.xlsx')
    wb = openpyxl.load_workbook(rec_file, data_only=True)

    # OS CKS sheet: outstanding AP checks
    os_cks = set()
    ws_cks = wb['OS CKS']
    for row in ws_cks.iter_rows(min_row=2, values_only=True):
        if row[2] is not None:  # Col C = Check#
            ck = str(row[2]).strip()
            if ck:
                os_cks.add(ck)

    # OS PR CKS sheet: outstanding payroll checks
    os_pr = set()
    ws_pr = wb['OS PR CKS']
    for row in ws_pr.iter_rows(min_row=2, values_only=True):
        if row[2] is not None:  # Col C = Check#
            ck = str(row[2]).strip()
            if ck:
                os_pr.add(ck)

    # OS DEPOSITS sheet: outstanding deposits (reference codes)
    os_deps = []
    ws_dep = wb['OS DEPOSITS']
    for row in ws_dep.iter_rows(min_row=2, values_only=True):
        if row[3] is not None and row[4] is not None:  # Col D=Ref, E=Amount
            os_deps.append({
                'ref': str(row[3]).strip(),
                'amount': float(row[4]),
            })

    wb.close()
    return os_cks, os_pr, os_deps


def main():
    print("=" * 70)
    print("ABR Honda Feb 2026 — Full Pipeline Test")
    print("=" * 70)

    # --- Step 1: Format Detection ---
    fmt = detect_bank_format(BANK_FILE)
    print(f"\nBank format detected: {fmt}")
    assert fmt == "BOFA_BAI", f"Expected BOFA_BAI, got {fmt}"

    # --- Step 2: Parse Bank Statement ---
    print("\nParsing Honda bank statement (BAI format)...")
    bank_txns = parse_bofa_bai_csv(BANK_FILE)
    print(f"  Bank transactions loaded: {len(bank_txns)}")

    # Separate reconciling items from matchable transactions
    reconciling = [t for t in bank_txns if t.is_reconciling_item]
    matchable = [t for t in bank_txns if not t.is_reconciling_item]
    sweeps = [t for t in reconciling if t.reconciling_type == 'SWEEP']
    securities = [t for t in reconciling if t.reconciling_type == 'SECURITIES']

    print(f"  Matchable transactions: {len(matchable)}")
    print(f"  Reconciling items excluded: {len(reconciling)}")
    print(f"    Sweep transfers (BAI 501): {len(sweeps)} (${sum(t.amount for t in sweeps):,.2f})")
    print(f"    Securities sold (BAI 233): {len(securities)} (${sum(t.amount for t in securities):,.2f})")

    # Sanity checks on matchable transactions only
    checks = [t for t in matchable if t.check_number]
    ach_debits = [t for t in matchable if t.type_code == 'ACH_DR']
    ach_credits = [t for t in matchable if t.type_code == 'ACH_CR']
    deposits = [t for t in matchable if t.type_code == 'DEP']
    debits = [t for t in matchable if t.amount < 0]
    credits = [t for t in matchable if t.amount > 0]

    print(f"\n  Matchable breakdown:")
    print(f"  Checks (BAI 475): {len(checks)}")
    print(f"  ACH debits: {len(ach_debits)}")
    print(f"  ACH credits: {len(ach_credits)}")
    print(f"  Deposits: {len(deposits)}")
    print(f"  Total debits: {len(debits)} (${sum(t.amount for t in debits):,.2f})")
    print(f"  Total credits: {len(credits)} (${sum(t.amount for t in credits):,.2f})")

    # --- Step 3: Parse DMS GL (Feb + Jan) ---
    print("\nParsing Honda Feb GL...")
    feb_dms = parse_dms_xlsx(FEB_GL_FILE)
    print(f"  Feb GL transactions: {len(feb_dms)}")

    print("Parsing Honda Jan GL (prior period)...")
    jan_dms = parse_dms_xlsx(JAN_GL_FILE)
    print(f"  Jan GL transactions: {len(jan_dms)}")

    # DMS breakdown
    chk_dms = [t for t in feb_dms if t.type_code == 'CHK']
    batch_dms = [t for t in feb_dms if t.type_code == 'BATCH']
    findep_dms = [t for t in feb_dms if t.type_code == 'FINDEP']
    print(f"  Feb CHK (SRC=6): {len(chk_dms)}")
    print(f"  Feb BATCH (SRC=5): {len(batch_dms)}")
    print(f"  Feb FINDEP (SRC=11): {len(findep_dms)}")

    # --- Step 3b: Parse January Outstanding Checks (test fixture) ---
    jan_outstanding = None
    if os.path.exists(JAN_REC_FILE):
        print(f"\nParsing January 2026 outstanding checks (test fixture)...")
        jan_outstanding = parse_outstanding_from_rec(JAN_REC_FILE)
        print(f"  Jan outstanding checks loaded: {len(jan_outstanding)}")
        jan_os_total = sum(t.amount for t in jan_outstanding)
        print(f"  Total outstanding amount: ${jan_os_total:,.2f}")
    else:
        print(f"\n  (No January outstanding checks fixture found at {JAN_REC_FILE})")

    # --- Step 4: Run Matching Pipeline ---
    print("\n" + "-" * 70)
    print("Running matching pipeline...")
    print("-" * 70)

    config = MatchConfig()
    results = run_full_reconciliation(bank_txns, feb_dms, config,
                                      prior_dms_txns=jan_dms,
                                      prior_outstanding_txns=jan_outstanding)

    # --- Step 5: Results Summary ---
    total_matches = len(results['all_matches'])
    one_to_one = len(results['one_to_one_matches'])
    prior_period = len(results['prior_period_matches'])
    cvr = len(results['cvr_matches'])
    splits = len(results['split_matches'])
    prior_outstanding = len(results.get('prior_outstanding_matches', []))
    unmatched_bank = results['unmatched_bank']
    unmatched_dms = results['unmatched_dms']
    reconciling_items = results.get('reconciling_items', [])
    reconciling_by_type = results.get('reconciling_by_type', {})

    print(f"\n{'RESULTS':^70}")
    print("=" * 70)
    print(f"  Total bank transactions:    {results['total_bank']}")
    print(f"  Reconciling items excluded: {results.get('total_reconciling', 0)}")
    print(f"  Matchable bank transactions: {results.get('total_bank_matchable', results['total_bank'])}")
    print(f"  Total DMS transactions:     {results['total_dms']}")
    print(f"  1:1 matches (current):      {one_to_one}")
    print(f"  Prior-period matches:        {prior_period}")
    print(f"  CVR many-to-one:             {cvr}")
    print(f"  Reverse split:               {splits}")
    print(f"  Prior outstanding matches:   {prior_outstanding}")
    print(f"  Total matches:               {total_matches}")
    print(f"  Unmatched bank:              {len(unmatched_bank)}")
    print(f"  Unmatched DMS:               {len(unmatched_dms)}")

    # Reconciling items detail
    if reconciling_items:
        print(f"\n{'RECONCILING ITEMS (excluded from matching)':^70}")
        print("-" * 70)
        for rtype, items in sorted(reconciling_by_type.items()):
            total_amt = sum(t.amount for t in items)
            print(f"  {rtype}: {len(items)} transactions, total ${total_amt:,.2f}")
            for t in items:
                print(f"    {t.transaction_date}  ${t.amount:>12,.2f}  {t.description[:50]}")
        reconc_net = sum(t.amount for t in reconciling_items)
        print(f"  Net reconciling items: ${reconc_net:,.2f}")

    # Confidence distribution
    high_conf = [m for m in results['all_matches'] if m.confidence_score >= 85]
    med_conf = [m for m in results['all_matches'] if 60 <= m.confidence_score < 85]
    low_conf = [m for m in results['all_matches'] if m.confidence_score < 60]
    print(f"\n  High confidence (85%+):      {len(high_conf)}")
    print(f"  Medium confidence (60-84%):  {len(med_conf)}")
    print(f"  Low confidence (<60%):       {len(low_conf)}")

    # Auto-accept rate (85%+ out of matchable bank txns)
    matchable_bank_count = results.get('total_bank_matchable', results['total_bank'])
    if matchable_bank_count > 0:
        auto_accept_rate = len(high_conf) / matchable_bank_count * 100
        bank_match_rate = (matchable_bank_count - len(unmatched_bank)) / matchable_bank_count * 100
        print(f"\n  Auto-accept rate:            {auto_accept_rate:.1f}%")
        print(f"  Bank match rate:             {bank_match_rate:.1f}%")

    # Net difference on matched pairs
    total_diff = sum(abs(m.amount_difference) for m in results['all_matches'])
    print(f"  Net amount difference:       ${total_diff:.2f}")

    # --- Step 6: Unmatched Bank Analysis ---
    print(f"\n{'UNMATCHED BANK TRANSACTIONS':^70}")
    print("-" * 70)
    unmatched_by_type = {}
    for t in unmatched_bank:
        unmatched_by_type.setdefault(t.type_code, []).append(t)

    for ttype, txns in sorted(unmatched_by_type.items()):
        total_amt = sum(t.amount for t in txns)
        print(f"  {ttype}: {len(txns)} transactions, total ${total_amt:,.2f}")

    # Show unmatched checks specifically
    unmatched_bank_checks = [t for t in unmatched_bank if t.check_number]
    if unmatched_bank_checks:
        print(f"\n  Unmatched bank checks: {len(unmatched_bank_checks)}")
        # Show first 20
        for t in sorted(unmatched_bank_checks, key=lambda x: x.check_number)[:20]:
            print(f"    Check #{t.check_number}  ${t.amount:>12,.2f}  {t.transaction_date}")
        if len(unmatched_bank_checks) > 20:
            print(f"    ... and {len(unmatched_bank_checks) - 20} more")

    # --- Step 7: Unmatched DMS Analysis ---
    print(f"\n{'UNMATCHED DMS TRANSACTIONS':^70}")
    print("-" * 70)
    unmatched_dms_by_type = {}
    for t in unmatched_dms:
        unmatched_dms_by_type.setdefault(t.type_code, []).append(t)

    for ttype, txns in sorted(unmatched_dms_by_type.items()):
        total_amt = sum(t.amount for t in txns)
        print(f"  {ttype}: {len(txns)} transactions, total ${total_amt:,.2f}")

    # --- Step 8: Validate Against Ground Truth ---
    print(f"\n{'GROUND TRUTH VALIDATION':^70}")
    print("=" * 70)

    os_cks, os_pr_cks, os_deps = load_outstanding_checks_from_rec()

    if os_cks:
        all_os_checks = os_cks | os_pr_cks
        print(f"\n  Outstanding checks in bank rec: {len(os_cks)} AP + {len(os_pr_cks)} PR = {len(all_os_checks)} total")

        # Bank checks that cleared (present in bank statement)
        bank_check_nums = {t.check_number for t in bank_txns if t.check_number}
        print(f"  Checks in bank statement: {len(bank_check_nums)}")

        # Checks that are on the bank statement AND on outstanding list = problem (should have been removed)
        cleared_but_outstanding = bank_check_nums & all_os_checks
        if cleared_but_outstanding:
            print(f"  WARNING: {len(cleared_but_outstanding)} checks appear on BOTH bank stmt and outstanding list")
            for ck in sorted(cleared_but_outstanding)[:10]:
                print(f"    Check #{ck}")
        else:
            print(f"  OK: No checks appear on both bank statement and outstanding list")

        # Our unmatched bank checks vs outstanding checks
        our_unmatched_check_nums = {t.check_number for t in unmatched_bank if t.check_number}
        print(f"\n  Our unmatched bank checks: {len(our_unmatched_check_nums)}")

        # Unmatched checks that ARE on the outstanding list (expected — these are prior-period checks
        # on the bank stmt that haven't been matched to GL because they were cut in a prior period
        # and aren't in the current or prior month GL)
        our_unmatched_on_os = our_unmatched_check_nums & all_os_checks
        our_unmatched_not_on_os = our_unmatched_check_nums - all_os_checks
        print(f"  Of those, on outstanding list: {len(our_unmatched_on_os)} (expected)")
        print(f"  Of those, NOT on outstanding list: {len(our_unmatched_not_on_os)} (investigate)")
        if our_unmatched_not_on_os:
            for ck in sorted(our_unmatched_not_on_os)[:10]:
                txn = next(t for t in unmatched_bank if t.check_number == ck)
                print(f"    Check #{ck}  ${txn.amount:>12,.2f}  {txn.transaction_date}")

    # --- Step 8b: Outstanding Checks Parser Validation ---
    print(f"\n{'OUTSTANDING CHECKS PARSER (parse_outstanding_from_rec)':^70}")
    print("-" * 70)

    os_txns = parse_outstanding_from_rec(REC_FILE)
    print(f"  Parsed {len(os_txns)} outstanding checks from bank rec OS CKS sheet")
    os_total = sum(t.amount for t in os_txns)
    print(f"  Total outstanding amount: ${os_total:,.2f}")
    print(f"  Sample (first 5):")
    for t in os_txns[:5]:
        print(f"    Check #{t.check_number}  ${t.amount:>12,.2f}  {t.description[:40]}")

    # Validate parser output matches the raw read used earlier
    os_txn_check_nums = {t.check_number for t in os_txns}
    if os_cks:
        # os_cks was read with int->str conversion; os_txns uses int->str too
        # Normalize: strip ".0" from floats that were read as strings
        os_cks_normalized = set()
        for ck in os_cks:
            try:
                os_cks_normalized.add(str(int(float(ck))))
            except (ValueError, TypeError):
                os_cks_normalized.add(ck)

        parser_vs_raw = os_txn_check_nums.symmetric_difference(os_cks_normalized)
        if parser_vs_raw:
            print(f"  WARNING: Parser vs raw read mismatch: {len(parser_vs_raw)} differences")
            for ck in sorted(parser_vs_raw)[:5]:
                print(f"    {ck}")
        else:
            print(f"  OK: Parser output matches raw read ({len(os_txn_check_nums)} checks)")

    # --- Step 8c: DMS Unmatched Checks vs Outstanding Checks Overlap ---
    print(f"\n{'DMS UNMATCHED vs OUTSTANDING CHECKS OVERLAP':^70}")
    print("-" * 70)

    # Our unmatched DMS checks (type CHK) should overlap with outstanding list
    # because outstanding checks = GL entries that haven't cleared the bank yet
    unmatched_dms_checks = [t for t in unmatched_dms if t.type_code == 'CHK'
                            and t.check_number]
    print(f"  Unmatched DMS checks (CHK type with check#): {len(unmatched_dms_checks)}")

    unmatched_dms_check_nums = set()
    for t in unmatched_dms_checks:
        # DMS check numbers may come from reference_number with trailing letters
        ck = t.check_number
        if not ck and t.reference_number:
            ck = t.reference_number.rstrip('ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz')
        if ck:
            unmatched_dms_check_nums.add(ck)

    print(f"  Unique DMS check numbers: {len(unmatched_dms_check_nums)}")

    dms_on_os = unmatched_dms_check_nums & os_txn_check_nums
    dms_not_on_os = unmatched_dms_check_nums - os_txn_check_nums
    print(f"  DMS checks found on outstanding list: {len(dms_on_os)}")
    print(f"  DMS checks NOT on outstanding list: {len(dms_not_on_os)}")

    if len(unmatched_dms_check_nums) > 0:
        overlap_pct = len(dms_on_os) / len(unmatched_dms_check_nums) * 100
        print(f"  Overlap rate: {overlap_pct:.1f}%")
        print(f"  (High overlap validates that our engine correctly identifies")
        print(f"   outstanding checks — they are in the GL but not on the bank stmt)")

    if dms_not_on_os:
        print(f"\n  DMS checks not on outstanding list (investigate):")
        for ck in sorted(dms_not_on_os)[:10]:
            txn = next((t for t in unmatched_dms_checks if t.check_number == ck), None)
            if txn:
                print(f"    Check #{ck}  ${txn.amount:>12,.2f}  {txn.transaction_date}  {txn.description[:30]}")

    # --- Step 8d: Prior Outstanding Match Results ---
    print(f"\n{'PRIOR OUTSTANDING MATCHING — RESULTS':^70}")
    print("-" * 70)
    if prior_outstanding:
        print(f"  Prior outstanding matches: {prior_outstanding}")
        print(f"  These bank checks cleared in Feb but were cut in a prior period.")
        print(f"  Matched against January 2026 outstanding checks list (test fixture).")
        print()
        for m in results['prior_outstanding_matches']:
            print(f"    Match #{m.match_id}: ${m.bank_amount:>12,.2f} "
                  f"conf={m.confidence_score:.0f}% {m.score_breakdown[:60]}")
    else:
        print("  No prior outstanding matches (fixture not loaded or no matches found).")

    if os_deps:
        print(f"\n  Outstanding deposits in bank rec: {len(os_deps)}")
        os_dep_refs = {d['ref'] for d in os_deps}
        os_dep_total = sum(d['amount'] for d in os_deps)
        print(f"  Outstanding deposit total: ${os_dep_total:,.2f}")
        for d in os_deps:
            print(f"    {d['ref']}: ${d['amount']:,.2f}")

        # Validate outstanding deposit protection:
        # All 6 outstanding deposits should be UNMATCHED (in DMS but not matched).
        # Use current-period match IDs only (prior-period matches use jan_dms which
        # has overlapping transaction IDs with feb_dms).
        current_matched_dms_ids = set()
        for m in results['one_to_one_matches']:
            current_matched_dms_ids.update(m.dms_transaction_ids)

        false_matched_count = 0
        correctly_unmatched = 0
        print(f"\n  Outstanding deposit protection check:")
        for dep in os_deps:
            amt = round(dep['amount'], 2)
            # Find BATCH DMS entries at this amount
            dms_at_amt = [t for t in feb_dms if t.type_code == 'BATCH'
                          and round(t.amount, 2) == amt]
            for t in dms_at_amt:
                if t.transaction_id in current_matched_dms_ids:
                    # Find which match consumed this DMS entry
                    culprit = next((m for m in results['one_to_one_matches']
                                    if t.transaction_id in m.dms_transaction_ids), None)
                    culprit_info = ""
                    if culprit:
                        culprit_info = (f" match#{culprit.match_id} "
                                        f"type={culprit.match_type} "
                                        f"conf={culprit.confidence_score}% "
                                        f"bd={culprit.score_breakdown[:60]}")
                    print(f"    FAIL: {dep['ref']} ${amt:,.2f} (DMS #{t.transaction_id} "
                          f"{t.transaction_date}) was falsely matched —{culprit_info}")
                    false_matched_count += 1
                else:
                    print(f"    OK:   {dep['ref']} ${amt:,.2f} (DMS #{t.transaction_id} "
                          f"{t.transaction_date}) correctly unmatched")
                    correctly_unmatched += 1
        if false_matched_count == 0:
            print(f"  PASS: All {correctly_unmatched} outstanding deposits correctly unmatched")
        else:
            print(f"  FAIL: {false_matched_count} outstanding deposits falsely matched!")

    # --- Step 9: Match quality spot check ---
    print(f"\n{'MATCH QUALITY SPOT CHECK':^70}")
    print("-" * 70)

    # Check matches — these should be high confidence
    check_matches = [m for m in results['all_matches']
                     if m.check_number_match is True]
    print(f"  Check# confirmed matches: {len(check_matches)}")
    if check_matches:
        avg_conf = sum(m.confidence_score for m in check_matches) / len(check_matches)
        print(f"  Average confidence: {avg_conf:.1f}%")

    # Show lowest-confidence matches for review
    all_sorted = sorted(results['all_matches'], key=lambda m: m.confidence_score)
    print(f"\n  Bottom 10 matches by confidence:")
    for m in all_sorted[:10]:
        bank_ids = ','.join(str(x) for x in m.bank_transaction_ids)
        print(f"    Match #{m.match_id}: {m.confidence_score:.1f}% "
              f"${m.bank_amount:>12,.2f} vs ${m.dms_amount:>12,.2f} "
              f"diff=${m.amount_difference:.2f} type={m.match_type}")

    print("\n" + "=" * 70)
    print("Done.")
    return results


if __name__ == '__main__':
    main()
