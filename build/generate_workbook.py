#!/usr/bin/env python3
"""
ABR Workbook Generator

Creates the Excel workbook (.xlsx) with all 9 sheets, formatting,
named ranges, conditional formatting, and default configuration.

Usage:
    python generate_workbook.py                    # Default location
    python generate_workbook.py --location JCC     # Specific location
    python generate_workbook.py --all              # All 7 locations
"""

import argparse
import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
    )
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    print("openpyxl is required. Install with: pip3 install openpyxl")
    sys.exit(1)

BUILD_DIR = os.path.dirname(os.path.abspath(__file__))
DIST_DIR = os.path.join(BUILD_DIR, '..', 'dist')
CONFIG_FILE = os.path.join(BUILD_DIR, 'build_config.json')


def load_config():
    with open(CONFIG_FILE, 'r') as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Style definitions
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style='medium', color='000000'),
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
)

DATA_FONT = Font(name='Calibri', size=11)
DATA_BORDER = Border(
    bottom=Side(style='thin', color='E0E0E0'),
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
)

TITLE_FONT = Font(name='Calibri', bold=True, size=18, color='1F4E79')
SUBTITLE_FONT = Font(name='Calibri', size=14, color='4472C4')
LABEL_FONT = Font(name='Calibri', bold=True, size=11, color='333333')
VALUE_FONT = Font(name='Calibri', size=12, color='1F4E79')

# Confidence band fills
GREEN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
RED_FILL = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')

FORMAT_MAP = {
    'currency': '#,##0.00',
    'percent': '0.0%',
    'date': 'MM/DD/YYYY',
    'datetime': 'MM/DD/YYYY HH:MM:SS',
    'integer': '0',
    'text': '@',
    'boolean': '@',
}


def apply_header_style(ws, col_count):
    """Apply header styling to row 1."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER


def create_data_sheet(wb, sheet_name, sheet_config):
    """Create a standard data sheet with headers and formatting."""
    ws = wb.create_sheet(title=sheet_name)

    columns = sheet_config.get('columns', [])
    if not columns:
        return ws

    # Write headers
    for i, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=i, value=col['header'])
        ws.column_dimensions[get_column_letter(i)].width = col.get('width', 12)

    apply_header_style(ws, len(columns))

    # Set number formats for columns
    for i, col in enumerate(columns, 1):
        fmt = col.get('format', 'text')
        if fmt in FORMAT_MAP:
            # Set format on a range of rows (pre-format for data entry)
            for row in range(2, 502):  # Pre-format 500 rows
                cell = ws.cell(row=row, column=i)
                cell.number_format = FORMAT_MAP[fmt]
                cell.font = DATA_FONT
                cell.border = DATA_BORDER

    # Freeze pane
    freeze = sheet_config.get('freeze_pane')
    if freeze:
        ws.freeze_panes = freeze

    # Tab color
    tab_color = sheet_config.get('tab_color')
    if tab_color:
        ws.sheet_properties.tabColor = tab_color

    return ws


def create_dashboard(wb, location_name, location_code):
    """Create the Dashboard sheet with layout and buttons."""
    ws = wb.create_sheet(title='Dashboard')
    ws.sheet_properties.tabColor = '1F4E79'

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 3

    # Title
    ws.merge_cells('B2:F2')
    title_cell = ws['B2']
    title_cell.value = 'ABR - Auto Bank Reconciliation'
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center')

    # Location
    ws.merge_cells('B3:F3')
    loc_cell = ws['B3']
    loc_cell.value = f'{location_name} ({location_code})'
    loc_cell.font = SUBTITLE_FONT
    loc_cell.alignment = Alignment(horizontal='center')

    # Separator
    for col in range(2, 7):
        ws.cell(row=4, column=col).border = Border(
            bottom=Side(style='medium', color='1F4E79'))

    # Workflow Steps
    ws['B6'] = 'RECONCILIATION WORKFLOW'
    ws['B6'].font = Font(name='Calibri', bold=True, size=13, color='1F4E79')

    steps = [
        ('Step 1:', 'Import Bank Statement', 'Click to import bank CSV file'),
        ('Step 2:', 'Import DMS Data', 'Click to import R&R DMS GL export'),
        ('Step 3:', 'Run Auto-Matching', 'Run confidence-based matching engine'),
        ('Step 4:', 'Review & Confirm', 'Review staged matches, accept/reject'),
        ('Step 5:', 'Finalize & Export', 'Export reconciliation report'),
    ]

    step_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

    for i, (step, name, desc) in enumerate(steps):
        row = 8 + i * 2
        ws.cell(row=row, column=2, value=step).font = LABEL_FONT
        ws.cell(row=row, column=3, value=name).font = Font(
            name='Calibri', size=11, color='1F4E79')
        ws.cell(row=row, column=4, value='[ NOT STARTED ]').font = Font(
            name='Calibri', size=10, color='808080')
        ws.cell(row=row, column=5, value=desc).font = Font(
            name='Calibri', size=9, color='808080', italic=True)
        for col in range(2, 7):
            ws.cell(row=row, column=col).fill = step_fill

    # Separator
    stat_row = 20
    for col in range(2, 7):
        ws.cell(row=stat_row - 1, column=col).border = Border(
            bottom=Side(style='medium', color='1F4E79'))

    # Statistics section
    ws.cell(row=stat_row, column=2, value='RECONCILIATION SUMMARY').font = Font(
        name='Calibri', bold=True, size=13, color='1F4E79')

    stats = [
        ('Total Bank Transactions:', 0),
        ('Total DMS Transactions:', 0),
        ('Matched (1:1):', 0),
        ('Matched (CVR/Split):', 0),
        ('Staged for Review:', 0),
        ('Unmatched Bank:', 0),
        ('Unmatched DMS:', 0),
        ('Match Rate:', '0.0%'),
    ]

    for i, (label, value) in enumerate(stats):
        row = stat_row + 2 + i
        ws.cell(row=row, column=2, value=label).font = LABEL_FONT
        val_cell = ws.cell(row=row, column=3, value=value)
        val_cell.font = VALUE_FONT
        val_cell.alignment = Alignment(horizontal='right')

    # Named ranges for stats
    stat_start_row = stat_row + 2
    stat_names = [
        'TotalBankItems', 'TotalDMSItems', 'MatchedOneToOne',
        'MatchedCVRSplit', 'StagedForReview', 'UnmatchedBank',
        'UnmatchedDMS', 'MatchRate',
    ]
    for i, name in enumerate(stat_names):
        cell_ref = f"Dashboard!$C${stat_start_row + i}"
        defn = DefinedName(name, attr_text=cell_ref)
        wb.defined_names.add(defn)

    # Session info
    info_row = stat_start_row + len(stats) + 2
    ws.cell(row=info_row, column=2, value='LAST SESSION').font = Font(
        name='Calibri', bold=True, size=11, color='808080')
    ws.cell(row=info_row + 1, column=2, value='Date:').font = LABEL_FONT
    ws.cell(row=info_row + 2, column=2, value='User:').font = LABEL_FONT
    ws.cell(row=info_row + 3, column=2, value='Current Month:').font = LABEL_FONT

    return ws


def create_config_sheet(wb, sheet_config, location_name, location_code, bank_type):
    """Create the Config sheet with default values."""
    ws = wb.create_sheet(title='Config')
    ws.sheet_properties.tabColor = '808080'

    columns = sheet_config.get('columns', [])
    for i, col in enumerate(columns, 1):
        ws.cell(row=1, column=i, value=col['header'])
        ws.column_dimensions[get_column_letter(i)].width = col.get('width', 12)

    apply_header_style(ws, len(columns))

    defaults = sheet_config.get('default_values', [])
    for i, cfg in enumerate(defaults, 2):
        ws.cell(row=i, column=1, value=cfg['setting']).font = LABEL_FONT
        val = cfg['value']

        # Override location-specific values
        if cfg['setting'] == 'LocationName':
            val = location_name
        elif cfg['setting'] == 'LocationCode':
            val = location_code
        elif cfg['setting'] == 'BankType':
            val = bank_type

        ws.cell(row=i, column=2, value=val).font = DATA_FONT
        ws.cell(row=i, column=3, value=cfg['description']).font = Font(
            name='Calibri', size=10, color='808080', italic=True)
        ws.cell(row=i, column=4, value=cfg['valid_range']).font = Font(
            name='Calibri', size=10, color='A0A0A0')

    return ws


def create_lookups_sheet(wb, sheet_config):
    """Create the Lookups sheet with reference data."""
    ws = wb.create_sheet(title='Lookups')
    ws.sheet_properties.tabColor = 'A5A5A5'

    # DMS Type Codes
    ws['A1'] = 'DMS Type Codes'
    ws['A1'].font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
    ws['A2'] = 'Code'
    ws['B2'] = 'Description'
    ws['A2'].font = HEADER_FONT
    ws['B2'].font = HEADER_FONT
    ws['A2'].fill = HEADER_FILL
    ws['B2'].fill = HEADER_FILL

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30

    type_codes = sheet_config.get('dms_type_codes', [])
    for i, tc in enumerate(type_codes, 3):
        ws.cell(row=i, column=1, value=tc['code']).font = DATA_FONT
        ws.cell(row=i, column=2, value=tc['description']).font = DATA_FONT

    # Check number patterns section
    ws['D1'] = 'Check Number Patterns'
    ws['D1'].font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
    ws['D2'] = 'Bank'
    ws['E2'] = 'Pattern'
    ws['D2'].font = HEADER_FONT
    ws['E2'].font = HEADER_FONT
    ws['D2'].fill = HEADER_FILL
    ws['E2'].fill = HEADER_FILL

    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 40

    patterns = [
        ('BOFA', 'CHECK #NNNN or CHK #NNNN in Description'),
        ('BOFA', 'CK NNNN in Description'),
        ('TRUIST', 'CHECK NNNN in Description'),
        ('TRUIST', 'Separate Check Number field (if present)'),
    ]
    for i, (bank, pattern) in enumerate(patterns, 3):
        ws.cell(row=i, column=4, value=bank).font = DATA_FONT
        ws.cell(row=i, column=5, value=pattern).font = DATA_FONT

    return ws


def add_conditional_formatting(ws, confidence_col, num_rows=500):
    """Add confidence-based conditional formatting to a sheet."""
    col_letter = get_column_letter(confidence_col)
    cell_range = f'{col_letter}2:{col_letter}{num_rows}'

    # Green for high confidence (>= 0.85)
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='greaterThanOrEqual', formula=['0.85'],
                   fill=GREEN_FILL))

    # Yellow for medium confidence (0.60 - 0.84)
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='between', formula=['0.60', '0.8499'],
                   fill=YELLOW_FILL))

    # Red for low confidence (< 0.60)
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='lessThan', formula=['0.60'],
                   fill=RED_FILL))


def generate_workbook(location_name: str, location_code: str, bank_type: str,
                      output_dir: str = None):
    """Generate a complete ABR workbook for a location."""
    config = load_config()
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets in order
    create_dashboard(wb, location_name, location_code)

    sheets_config = config['sheets']

    # Data sheets
    for sheet_name in ['BankData', 'DMSData']:
        ws = create_data_sheet(wb, sheet_name, sheets_config[sheet_name])
        # Add confidence conditional formatting
        col_count = len(sheets_config[sheet_name]['columns'])
        conf_col = next((i + 1 for i, c in enumerate(sheets_config[sheet_name]['columns'])
                         if c['header'] == 'Confidence'), None)
        if conf_col:
            add_conditional_formatting(ws, conf_col)

    # StagedMatches with conditional formatting
    ws = create_data_sheet(wb, 'StagedMatches', sheets_config['StagedMatches'])
    add_conditional_formatting(ws, 3)  # Column C = Confidence

    # Reconciled
    create_data_sheet(wb, 'Reconciled', sheets_config['Reconciled'])

    # Outstanding
    create_data_sheet(wb, 'Outstanding', sheets_config['Outstanding'])

    # AuditLog
    create_data_sheet(wb, 'AuditLog', sheets_config['AuditLog'])

    # Config
    create_config_sheet(wb, sheets_config['Config'], location_name, location_code, bank_type)

    # Lookups
    create_lookups_sheet(wb, sheets_config['Lookups'])

    # Save
    if output_dir is None:
        output_dir = DIST_DIR

    os.makedirs(output_dir, exist_ok=True)
    filename = f'ABR_{location_code}.xlsx'
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"  Created: {filepath}")
    return filepath


def main():
    parser = argparse.ArgumentParser(description='Generate ABR workbook(s)')
    parser.add_argument('--location', type=str, help='Location code (e.g., JCC)')
    parser.add_argument('--all', action='store_true', help='Generate all 7 locations')
    parser.add_argument('--output', type=str, help='Output directory')
    args = parser.parse_args()

    config = load_config()
    locations = config['locations']

    if args.all:
        print("Generating workbooks for all locations...")
        for loc in locations:
            generate_workbook(loc['name'], loc['code'], loc['bank'], args.output)
        print(f"\nGenerated {len(locations)} workbooks.")
    elif args.location:
        loc = next((l for l in locations if l['code'] == args.location.upper()), None)
        if loc is None:
            print(f"Unknown location: {args.location}")
            print(f"Valid codes: {', '.join(l['code'] for l in locations)}")
            sys.exit(1)
        generate_workbook(loc['name'], loc['code'], loc['bank'], args.output)
    else:
        # Default: generate first location as demo
        loc = locations[0]
        print(f"Generating demo workbook for {loc['name']}...")
        generate_workbook(loc['name'], loc['code'], loc['bank'], args.output)
        print("\nUse --all to generate all 7 locations.")


if __name__ == '__main__':
    main()
